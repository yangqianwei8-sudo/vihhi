from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Count, Sum, Q, F, Max
from django.core.paginator import Paginator
from django.urls import reverse, NoReverseMatch
from django.utils import timezone
from django.forms import inlineformset_factory
from django.http import HttpResponse, JsonResponse
from datetime import timedelta, datetime
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from backend.apps.system_management.services import get_user_permission_codes
from backend.core.views import HOME_NAV_STRUCTURE, _permission_granted as core_permission_granted, _build_full_top_nav
from backend.apps.financial_management.models import (
    AccountSubject, Voucher, VoucherEntry,
    Ledger, Budget, Invoice, FundFlow,
    FinancialReport, ReceivableAccount, PayableAccount,
)
from .forms import (
    AccountSubjectForm, VoucherForm, VoucherEntryForm, BudgetForm, InvoiceForm, FundFlowForm
)


def _permission_granted(required_code, user_permissions: set) -> bool:
    """检查权限"""
    if not required_code:
        return True
    if '__all__' in user_permissions:
        return True
    return required_code in user_permissions


def _generate_voucher_number(voucher_date=None):
    """
    自动生成凭证字号
    
    Args:
        voucher_date: 凭证日期，如果为None则使用当前日期
    
    Returns:
        str: 生成的凭证字号，格式为 VOUCHER-YYYY-NNNN
    """
    if voucher_date:
        current_year = voucher_date.year
    else:
        current_year = timezone.now().year
    
    # 查找当前年度最大的凭证号
    max_voucher = Voucher.objects.filter(
        voucher_number__startswith=f'VOUCHER-{current_year}-'
    ).aggregate(max_num=Max('voucher_number'))['max_num']
    
    if max_voucher:
        try:
            # 提取序号部分
            seq = int(max_voucher.split('-')[-1]) + 1
        except (ValueError, IndexError):
            seq = 1
    else:
        seq = 1
    
    return f'VOUCHER-{current_year}-{seq:04d}'


def _update_budget_from_fund_flow(fund_flow, is_create=True, old_amount=None):
    """
    根据资金流水自动更新预算使用金额
    
    Args:
        fund_flow: 资金流水对象
        is_create: 是否为创建操作（True=创建，False=更新或删除）
        old_amount: 旧金额（更新或删除时需要）
    """
    try:
        # 只有支出类型的流水才更新预算
        if fund_flow.flow_type != 'expense':
            return
        
        # 查找相关预算（根据日期范围、状态）
        budgets = Budget.objects.filter(
            status__in=['approved', 'executing'],
            start_date__lte=fund_flow.flow_date,
            end_date__gte=fund_flow.flow_date,
        )
        
        # 如果有关联项目，可以通过项目查找相关预算
        # 优先匹配关联项目的预算
        if fund_flow.project:
            project_budgets = budgets.filter(project=fund_flow.project)
            if project_budgets.exists():
                budgets = project_budgets
        
        # 如果资金流水关联了凭证，可以通过凭证分录的会计科目匹配预算科目
        if fund_flow.voucher and fund_flow.voucher.entries.exists():
            # 获取凭证中支出相关的会计科目
            expense_subjects = set()
            for entry in fund_flow.voucher.entries.all():
                if entry.debit_amount > 0:  # 借方金额表示支出
                    expense_subjects.add(entry.account_subject)
            
            # 如果预算指定了会计科目，优先匹配
            if expense_subjects:
                subject_budgets = budgets.filter(account_subject__in=expense_subjects)
                if subject_budgets.exists():
                    budgets = subject_budgets
        
        # 更新预算使用金额
        if is_create:
            # 创建：增加使用金额
            amount_to_add = fund_flow.amount
            for budget in budgets:
                budget.used_amount += amount_to_add
                budget.remaining_amount = budget.budget_amount - budget.used_amount
                # 如果剩余金额为0或负数，更新状态
                if budget.remaining_amount <= 0:
                    budget.status = 'completed'
                budget.save()
        else:
            # 更新或删除：计算金额差异
            if old_amount is not None:
                amount_diff = fund_flow.amount - old_amount
                for budget in budgets:
                    budget.used_amount += amount_diff
                    budget.remaining_amount = budget.budget_amount - budget.used_amount
                    # 如果剩余金额为0或负数，更新状态
                    if budget.remaining_amount <= 0:
                        budget.status = 'completed'
                    elif budget.remaining_amount > 0 and budget.status == 'completed':
                        budget.status = 'executing'
                    budget.save()
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('更新预算失败: %s', str(e))
        # 不影响资金流水的保存，只记录日志


# 使用统一的顶部导航菜单生成函数
from backend.core.views import _build_full_top_nav


def _context(page_title, page_icon, description, summary_cards=None, sections=None, request=None, use_financial_nav=False):
    """构建页面上下文"""
    context = {
        "page_title": page_title,
        "page_icon": page_icon,
        "description": description,
        "summary_cards": summary_cards or [],
        "sections": sections or [],
    }
    
    try:
        if request and request.user.is_authenticated:
            permission_set = get_user_permission_codes(request.user)
            # 统一使用全局系统主菜单（与客户管理模块保持一致）
            context['full_top_nav'] = _build_full_top_nav(permission_set, request.user)
            if use_financial_nav:
                context['financial_menu'] = _build_financial_sidebar_nav(permission_set, request.path)
                context['module_sidebar_nav'] = context['financial_menu']  # 兼容模板中的变量名
        else:
            context['full_top_nav'] = []
            context['financial_menu'] = []
            context['module_sidebar_nav'] = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'构建页面上下文错误: {str(e)}', exc_info=True)
        context['full_top_nav'] = []
        context['financial_menu'] = []
        context['module_sidebar_nav'] = []
    
    # 为所有可能的侧边栏变量设置默认值，避免模板错误
    # 这些变量可能在其他模块的模板中被引用
    context.setdefault('plan_menu', [])
    context.setdefault('delivery_sidebar_nav', [])
    context.setdefault('customer_menu', [])
    context.setdefault('production_sidebar_nav', [])
    context.setdefault('personnel_sidebar_nav', [])
    context.setdefault('sidebar_menu', [])
    context.setdefault('litigation_sidebar_nav', [])
    context.setdefault('archive_sidebar_nav', [])
    context.setdefault('production_management_menu', [])
    
    return context


def _build_financial_top_nav(permission_set):
    """
    生成财务管理专用的顶部导航菜单 - 已废弃
    
    注意：此函数已不再使用，系统现在统一使用 _build_full_top_nav 生成全局系统主菜单。
    保留此函数仅用于历史参考，可以安全删除。
    """
    # 定义财务管理功能模块（从左到右的顺序）
    financial_modules = [
        {
            'label': '会计科目',
            'url_name': 'finance_pages:account_subject_management',
            'permission': 'financial_management.account.view',
            'icon': '📊',
        },
        {
            'label': '凭证管理',
            'url_name': 'finance_pages:voucher_management',
            'permission': 'financial_management.voucher.view',
            'icon': '📝',
        },
        {
            'label': '账簿管理',
            'url_name': 'finance_pages:ledger_management',
            'permission': 'financial_management.ledger.view',
            'icon': '📖',
        },
        {
            'label': '预算管理',
            'url_name': 'finance_pages:budget_management',
            'permission': 'financial_management.budget.view',
            'icon': '💰',
        },
        {
            'label': '发票管理',
            'url_name': 'finance_pages:invoice_management',
            'permission': 'financial_management.invoice.view',
            'icon': '🧾',
        },
        {
            'label': '资金流水',
            'url_name': 'finance_pages:fund_flow_management',
            'permission': 'financial_management.fund_flow.view',
            'icon': '💳',
        },
        {
            'label': '财务报表',
            'url_name': 'finance_pages:report_management',
            'permission': 'financial_management.report.view',
            'icon': '📊',
        },
        {
            'label': '应收账款',
            'url_name': 'finance_pages:receivable_management',
            'permission': 'financial_management.receivable.view',
            'icon': '💰',
        },
        {
            'label': '应付账款',
            'url_name': 'finance_pages:payable_management',
            'permission': 'financial_management.payable.view',
            'icon': '💸',
        },
    ]
    
    # 过滤有权限的模块
    nav_items = []
    for module in financial_modules:
        if _permission_granted(module['permission'], permission_set):
            try:
                url = reverse(module['url_name'])
            except NoReverseMatch:
                url = '#'
            nav_items.append({
                'label': module['label'],
                'url': url,
                'icon': module.get('icon', ''),
            })
    
    return nav_items


def _build_financial_sidebar_nav(permission_set, request_path=None, active_id=None):
    """生成财务管理模块的左侧菜单导航（使用计划管理格式）
    
    Args:
        permission_set: 用户权限集合
        request_path: 当前请求路径，用于判断激活状态
        active_id: 当前激活的菜单项ID（可选）
    
    Returns:
        list: 分组菜单项列表，格式与计划管理一致
    """
    from django.urls import reverse, NoReverseMatch
    
    # 定义财务管理菜单结构（分组格式，与计划管理一致）
    FINANCIAL_MENU_STRUCTURE = [
        {
            'id': 'financial_basic',
            'label': '基础管理',
            'icon': '📊',
            'permission': 'financial_management.account.view',
            'children': [
                {
                    'id': 'financial_home',
                    'label': '财务管理首页',
                    'icon': '💵',
                    'url_name': 'finance_pages:financial_home',
                    'permission': None,
                    'path_keywords': ['financial_home', 'financial'],
                },
                {
                    'id': 'account_subject',
                    'label': '会计科目',
                    'icon': '📊',
                    'url_name': 'finance_pages:account_subject_management',
                    'permission': 'financial_management.account.view',
                    'path_keywords': ['account', 'accounts'],
                },
                {
                    'id': 'voucher',
                    'label': '凭证管理',
                    'icon': '📝',
                    'url_name': 'finance_pages:voucher_management',
                    'permission': 'financial_management.voucher.view',
                    'path_keywords': ['voucher', 'vouchers'],
                },
            ]
        },
        {
            'id': 'financial_ledger',
            'label': '账簿管理',
            'icon': '📖',
            'permission': 'financial_management.ledger.view',
            'children': [
                {
                    'id': 'ledger',
                    'label': '总账',
                    'icon': '📖',
                    'url_name': 'finance_pages:ledger_management',
                    'permission': 'financial_management.ledger.view',
                    'path_keywords': ['ledger', 'ledgers'],
                },
                {
                    'id': 'subsidiary_ledger',
                    'label': '明细账',
                    'icon': '📋',
                    'url_name': 'finance_pages:subsidiary_ledger',
                    'permission': 'financial_management.ledger.view',
                    'path_keywords': ['subsidiary'],
                },
                {
                    'id': 'balance_sheet',
                    'label': '科目余额表',
                    'icon': '📊',
                    'url_name': 'finance_pages:account_balance_sheet',
                    'permission': 'financial_management.ledger.view',
                    'path_keywords': ['balance-sheet'],
                },
                {
                    'id': 'trial_balance',
                    'label': '试算平衡表',
                    'icon': '⚖️',
                    'url_name': 'finance_pages:trial_balance',
                    'permission': 'financial_management.ledger.view',
                    'path_keywords': ['trial-balance'],
                },
            ]
        },
        {
            'id': 'financial_budget',
            'label': '预算与资金',
            'icon': '💰',
            'permission': 'financial_management.budget.view',
            'children': [
                {
                    'id': 'budget',
                    'label': '预算管理',
                    'icon': '💰',
                    'url_name': 'finance_pages:budget_management',
                    'permission': 'financial_management.budget.view',
                    'path_keywords': ['budget', 'budgets'],
                },
                {
                    'id': 'fund_flow',
                    'label': '资金流水',
                    'icon': '💳',
                    'url_name': 'finance_pages:fund_flow_management',
                    'permission': 'financial_management.fund_flow.view',
                    'path_keywords': ['fund-flow', 'fund_flow'],
                },
            ]
        },
        {
            'id': 'financial_invoice',
            'label': '发票与账款',
            'icon': '🧾',
            'permission': 'financial_management.invoice.view',
            'children': [
                {
                    'id': 'invoice',
                    'label': '发票管理',
                    'icon': '🧾',
                    'url_name': 'finance_pages:invoice_management',
                    'permission': 'financial_management.invoice.view',
                    'path_keywords': ['invoice', 'invoices'],
                },
                {
                    'id': 'receivable',
                    'label': '应收账款',
                    'icon': '💰',
                    'url_name': 'finance_pages:receivable_management',
                    'permission': 'financial_management.receivable.view',
                    'path_keywords': ['receivable', 'receivables'],
                },
                {
                    'id': 'payable',
                    'label': '应付账款',
                    'icon': '💸',
                    'url_name': 'finance_pages:payable_management',
                    'permission': 'financial_management.payable.view',
                    'path_keywords': ['payable', 'payables'],
                },
            ]
        },
        {
            'id': 'financial_report',
            'label': '财务报表',
            'icon': '📈',
            'permission': 'financial_management.report.view',
            'children': [
                {
                    'id': 'report',
                    'label': '财务报表',
                    'icon': '📊',
                    'url_name': 'finance_pages:report_management',
                    'permission': 'financial_management.report.view',
                    'path_keywords': ['report', 'reports', 'balance-sheet', 'income-statement', 'cash-flow'],
                },
            ]
        },
    ]
    
    menu = []
    
    for menu_group in FINANCIAL_MENU_STRUCTURE:
        # 检查父菜单权限
        permission = menu_group.get('permission')
        if permission and not _permission_granted(permission, permission_set):
            continue
        
        # 处理子菜单
        children = []
        for child in menu_group.get('children', []):
            # 检查子菜单权限
            child_permission = child.get('permission')
            if child_permission and not _permission_granted(child_permission, permission_set):
                continue
            
            # 获取URL
            url_name = child.get('url_name')
            url = '#'
            if url_name:
                try:
                    url = reverse(url_name)
                except NoReverseMatch:
                    url = '#'
            
            # 判断是否激活
            is_active = False
            if active_id:
                is_active = child.get('id') == active_id
            elif request_path:
                # 特殊处理首页
                if child.get('id') == 'financial_home':
                    try:
                        home_url = reverse('finance_pages:financial_home')
                        try:
                            home_url2 = reverse('finance_pages:financial_management_home')
                        except NoReverseMatch:
                            home_url2 = None
                        is_active = (
                            request_path == home_url or
                            (home_url2 and request_path == home_url2) or
                            request_path == '/financial/' or
                            request_path == '/financial/home/'
                        )
                    except NoReverseMatch:
                        pass
                if not is_active:
                    for keyword in child.get('path_keywords', []):
                        if keyword in request_path:
                            is_active = True
                            break
            
            children.append({
                'id': child.get('id'),
                'label': child.get('label'),
                'icon': child.get('icon'),
                'url': url,
                'active': is_active,
            })
        
        # 如果父菜单没有可见的子菜单，跳过
        if not children:
            continue
        
        # 判断父菜单是否激活（任意子菜单激活则父菜单激活）
        has_active_child = any(child.get('active') for child in children)
        
        menu.append({
            'id': menu_group.get('id'),
            'label': menu_group.get('label'),
            'icon': menu_group.get('icon'),
            'active': has_active_child,
            'expanded': has_active_child,  # 如果有激活项，默认展开
            'children': children,
        })
    
    return menu


def _format_user_display(user, default='—'):
    """格式化用户显示名称"""
    if not user:
        return default
    if hasattr(user, 'get_full_name') and user.get_full_name():
        return user.get_full_name()
    return user.username if hasattr(user, 'username') else str(user)


@login_required
def financial_home(request):
    """财务管理首页 - 数据展示中心"""
    permission_codes = get_user_permission_codes(request.user)
    now = timezone.now()
    today = now.date()
    this_month_start = today.replace(day=1)
    seven_days_ago = today - timedelta(days=7)
    
    context = {}
    
    try:
        # ========== 核心指标卡片 ==========
        core_cards = []
        
        # 会计科目统计
        total_accounts = AccountSubject.objects.filter(is_active=True).count()
        accounts_by_type = AccountSubject.objects.filter(is_active=True).values('subject_type').annotate(count=Count('id'))
        
        # 凭证管理统计
        all_vouchers = Voucher.objects.all()
        total_vouchers = all_vouchers.count()
        pending_vouchers = all_vouchers.filter(status='submitted').count()
        approved_vouchers = all_vouchers.filter(status='approved').count()
        posted_vouchers = all_vouchers.filter(status='posted').count()
        this_month_vouchers = all_vouchers.filter(voucher_date__gte=this_month_start).count()
        
        # 账簿管理统计
        current_year = today.year
        current_month = today.month
        ledger_entries = Ledger.objects.filter(
            period_year=current_year,
            period_month=current_month
        ).count()
        
        # 预算管理统计
        all_budgets = Budget.objects.all()
        executing_budgets = all_budgets.filter(status='executing').count()
        total_budget = all_budgets.filter(status='executing').aggregate(
            total=Sum('budget_amount')
        )['total'] or Decimal('0')
        
        # 发票管理统计
        all_invoices = Invoice.objects.all()
        total_invoices = all_invoices.count()
        unverified_invoices = all_invoices.filter(status='issued').count()
        this_month_invoices = all_invoices.filter(invoice_date__gte=this_month_start).count()
        
        # 资金流水统计
        all_fund_flows = FundFlow.objects.all()
        this_month_flows = all_fund_flows.filter(flow_date__gte=this_month_start).count()
        this_month_income = all_fund_flows.filter(
            flow_date__gte=this_month_start,
            flow_type='income'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        this_month_expense = all_fund_flows.filter(
            flow_date__gte=this_month_start,
            flow_type='expense'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # 卡片1：会计科目
        core_cards.append({
            'label': '会计科目',
            'icon': '📊',
            'value': str(total_accounts),
            'subvalue': f'启用科目总数',
            'url': reverse('finance_pages:account_subject_management'),
            'variant': 'secondary'
        })
        
        # 卡片2：待审核凭证
        core_cards.append({
            'label': '待审核凭证',
            'icon': '📝',
            'value': str(pending_vouchers),
            'subvalue': f'本月凭证 {this_month_vouchers} 张',
            'url': reverse('finance_pages:voucher_management') + '?status=submitted',
            'variant': 'dark' if pending_vouchers > 0 else 'secondary'
        })
        
        # 卡片3：已过账凭证
        core_cards.append({
            'label': '已过账凭证',
            'icon': '✅',
            'value': str(posted_vouchers),
            'subvalue': f'凭证总数 {total_vouchers}',
            'url': reverse('finance_pages:voucher_management') + '?status=posted',
            'variant': 'secondary'
        })
        
        # 卡片4：预算管理
        core_cards.append({
            'label': '预算管理',
            'icon': '💰',
            'value': str(executing_budgets),
            'subvalue': f'执行中预算 · 总额 ¥{total_budget:,.0f}',
            'url': reverse('finance_pages:budget_management'),
            'variant': 'secondary'
        })
        
        # 卡片5：发票管理
        core_cards.append({
            'label': '发票管理',
            'icon': '🧾',
            'value': str(total_invoices),
            'subvalue': f'待认证 {unverified_invoices} | 本月 {this_month_invoices} 张',
            'url': reverse('finance_pages:invoice_management'),
            'variant': 'dark' if unverified_invoices > 0 else 'secondary'
        })
        
        # 卡片6：资金流水
        core_cards.append({
            'label': '资金流水',
            'icon': '💳',
            'value': str(this_month_flows),
            'subvalue': f'本月收入 ¥{this_month_income:,.0f} | 支出 ¥{this_month_expense:,.0f}',
            'url': reverse('finance_pages:fund_flow_management'),
            'variant': 'secondary'
        })
        
        context['core_cards'] = core_cards
        
        # ========== 风险预警 ==========
        risk_warnings = []
        
        # 待审核凭证（超过3天）
        stale_vouchers = all_vouchers.filter(
            status='submitted',
            created_time__lt=timezone.make_aware(datetime.combine(seven_days_ago, datetime.min.time()))
        ).select_related('preparer')[:5]
        
        for voucher in stale_vouchers:
            days_since_create = (today - voucher.created_time.date()).days
            preparer_name = _format_user_display(voucher.preparer) if voucher.preparer else '未知'
            risk_warnings.append({
                'type': 'voucher',
                'title': f'凭证号：{voucher.voucher_number}',
                'responsible': preparer_name,
                'days': days_since_create,
                'url': reverse('finance_pages:voucher_detail', args=[voucher.id])
            })
        
        # 预算超支风险
        over_budget = all_budgets.filter(
            status='executing',
            remaining_amount__lt=0
        ).select_related('created_by')[:5]
        
        for budget in over_budget:
            over_amount = abs(budget.remaining_amount)
            creator_name = _format_user_display(budget.created_by) if budget.created_by else '未知'
            risk_warnings.append({
                'type': 'budget',
                'title': f'{budget.name} - 超支 ¥{over_amount:,.2f}',
                'responsible': creator_name,
                'days': 0,
                'url': reverse('finance_pages:budget_detail', args=[budget.id])
            })
        
        context['risk_warnings'] = risk_warnings[:5]
        context['stale_vouchers_count'] = stale_vouchers.count()
        context['over_budget_count'] = over_budget.count()
        
        # ========== 待办事项 ==========
        todo_items = []
        
        # 待审核凭证
        pending_voucher_list = all_vouchers.filter(status='submitted').select_related('preparer')[:5]
        for voucher in pending_voucher_list:
            preparer_name = _format_user_display(voucher.preparer) if voucher.preparer else '未知'
            todo_items.append({
                'type': 'voucher',
                'title': f'凭证号：{voucher.voucher_number}',
                'voucher_number': voucher.voucher_number,
                'responsible': preparer_name,
                'url': reverse('finance_pages:voucher_detail', args=[voucher.id])
            })
        
        # 待认证发票
        unverified_invoice_list = all_invoices.filter(status='issued').select_related('created_by')[:5]
        for invoice in unverified_invoice_list:
            creator_name = _format_user_display(invoice.created_by) if invoice.created_by else '未知'
            todo_items.append({
                'type': 'invoice',
                'title': f'发票号：{invoice.invoice_number}',
                'invoice_number': invoice.invoice_number,
                'responsible': creator_name,
                'url': reverse('finance_pages:invoice_detail', args=[invoice.id])
            })
        
        context['todo_items'] = todo_items[:10]
        context['pending_approval_count'] = pending_vouchers + unverified_invoices
        context['todo_summary_url'] = reverse('finance_pages:voucher_management') + '?status=submitted'
        
        # ========== 我的工作 ==========
        my_work = {}
        
        # 我创建的凭证
        my_vouchers = all_vouchers.filter(preparer=request.user).order_by('-created_time')[:3]
        my_work['my_vouchers'] = [{
            'title': voucher.voucher_number,
            'status': voucher.get_status_display(),
            'url': reverse('finance_pages:voucher_detail', args=[voucher.id])
        } for voucher in my_vouchers]
        my_work['my_vouchers_count'] = all_vouchers.filter(preparer=request.user).count()
        
        # 我创建的发票
        my_invoices = all_invoices.filter(created_by=request.user).order_by('-created_time')[:3]
        my_work['my_invoices'] = [{
            'title': invoice.invoice_number,
            'status': invoice.get_status_display(),
            'url': reverse('finance_pages:invoice_detail', args=[invoice.id])
        } for invoice in my_invoices]
        my_work['my_invoices_count'] = all_invoices.filter(created_by=request.user).count()
        
        my_work['summary_url'] = reverse('finance_pages:voucher_management')
        
        context['my_work'] = my_work
        
        # ========== 最近活动 ==========
        recent_activities = {}
        
        # 最近创建的凭证
        recent_vouchers = all_vouchers.select_related('preparer').order_by('-created_time')[:5]
        recent_activities['recent_vouchers'] = [{
            'title': voucher.voucher_number,
            'creator': _format_user_display(voucher.preparer),
            'time': voucher.created_time,
            'url': reverse('finance_pages:voucher_detail', args=[voucher.id])
        } for voucher in recent_vouchers]
        
        # 最近创建的资金流水
        recent_fund_flows = all_fund_flows.select_related('created_by').order_by('-flow_date')[:5]
        recent_activities['recent_fund_flows'] = [{
            'title': f'资金流水：{fund_flow.get_flow_type_display()} ¥{fund_flow.amount:,.2f}',
            'creator': _format_user_display(fund_flow.created_by) if hasattr(fund_flow, 'created_by') and fund_flow.created_by else '未知',
            'time': fund_flow.created_time if hasattr(fund_flow, 'created_time') and fund_flow.created_time else fund_flow.flow_date,
            'url': reverse('finance_pages:fund_flow_detail', args=[fund_flow.id])
        } for fund_flow in recent_fund_flows]
        
        context['recent_activities'] = recent_activities
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取财务管理统计数据失败: %s', str(e))
        context.setdefault('core_cards', [])
        context.setdefault('risk_warnings', [])
        context.setdefault('todo_items', [])
        context.setdefault('my_work', {})
        context.setdefault('recent_activities', {})
    
    # 顶部操作栏
    top_actions = []
    if _permission_granted('financial_management.account.create', permission_codes):
        try:
            top_actions.append({
                'label': '新增会计科目',
                'url': reverse('finance_pages:account_subject_create'),
                'icon': '➕'
            })
        except Exception:
            pass
    
    if _permission_granted('financial_management.voucher.create', permission_codes):
        try:
            top_actions.append({
                'label': '创建凭证',
                'url': reverse('finance_pages:voucher_create'),
                'icon': '📝'
            })
        except Exception:
            pass
    
    context['top_actions'] = top_actions
    
    # 构建上下文
    page_context = _context(
        "财务管理",
        "💵",
        "数据展示中心 - 集中展示财务关键指标、状态与风险",
        request=request,
        use_financial_nav=True
    )
    
    # 设置侧边栏导航
    financial_sidebar_nav = _build_financial_sidebar_nav(permission_codes, request.path, active_id='financial_home')
    page_context['financial_menu'] = financial_sidebar_nav
    page_context['module_sidebar_nav'] = financial_sidebar_nav
    page_context['sidebar_title'] = '财务管理'
    page_context['sidebar_subtitle'] = 'Financial Management'
    
    # 合并所有数据
    page_context.update(context)
    
    return render(request, "financial_management/home.html", page_context)


@login_required
def financial_statistics(request):
    """财务统计仪表板"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.report.view', permission_codes):
        messages.error(request, '您没有权限查看财务统计')
        return redirect('finance_pages:financial_home')
    
    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    this_year_start = today.replace(month=1, day=1)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)
    
    # 获取统计参数
    period = request.GET.get('period', 'month')  # month, quarter, year
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month)) if period == 'month' else None
    
    # 确定统计时间范围
    if period == 'month' and month:
        period_start = today.replace(year=year, month=month, day=1)
        if month == 12:
            period_end = today.replace(year=year+1, month=1, day=1) - timedelta(days=1)
        else:
            period_end = today.replace(year=year, month=month+1, day=1) - timedelta(days=1)
    elif period == 'quarter':
        quarter = int(request.GET.get('quarter', (today.month - 1) // 3 + 1))
        period_start = today.replace(year=year, month=(quarter-1)*3+1, day=1)
        if quarter == 4:
            period_end = today.replace(year=year+1, month=1, day=1) - timedelta(days=1)
        else:
            period_end = today.replace(year=year, month=quarter*3+1, day=1) - timedelta(days=1)
    else:  # year
        period_start = today.replace(year=year, month=1, day=1)
        period_end = today.replace(year=year, month=12, day=31)
    
    statistics = {}
    
    try:
        # 凭证统计
        vouchers = Voucher.objects.filter(voucher_date__gte=period_start, voucher_date__lte=period_end)
        statistics['vouchers'] = {
            'total': vouchers.count(),
            'draft': vouchers.filter(status='draft').count(),
            'submitted': vouchers.filter(status='submitted').count(),
            'approved': vouchers.filter(status='approved').count(),
            'posted': vouchers.filter(status='posted').count(),
            'total_debit': vouchers.aggregate(total=Sum('total_debit'))['total'] or Decimal('0'),
            'total_credit': vouchers.aggregate(total=Sum('total_credit'))['total'] or Decimal('0'),
        }
        
        # 资金流水统计
        fund_flows = FundFlow.objects.filter(flow_date__gte=period_start, flow_date__lte=period_end)
        statistics['fund_flows'] = {
            'total': fund_flows.count(),
            'income': fund_flows.filter(flow_type='income').aggregate(total=Sum('amount'))['total'] or Decimal('0'),
            'expense': fund_flows.filter(flow_type='expense').aggregate(total=Sum('amount'))['total'] or Decimal('0'),
            'transfer': fund_flows.filter(flow_type='transfer').aggregate(total=Sum('amount'))['total'] or Decimal('0'),
            'net': (fund_flows.filter(flow_type='income').aggregate(total=Sum('amount'))['total'] or Decimal('0')) - 
                   (fund_flows.filter(flow_type='expense').aggregate(total=Sum('amount'))['total'] or Decimal('0')),
        }
        
        # 发票统计
        invoices = Invoice.objects.filter(invoice_date__gte=period_start, invoice_date__lte=period_end)
        statistics['invoices'] = {
            'total': invoices.count(),
            'input': invoices.filter(invoice_type='input').aggregate(total=Sum('total_amount'))['total'] or Decimal('0'),
            'output': invoices.filter(invoice_type='output').aggregate(total=Sum('total_amount'))['total'] or Decimal('0'),
            'issued': invoices.filter(status='issued').count(),
            'verified': invoices.filter(status='verified').count(),
        }
        
        # 应收账款统计
        receivables = ReceivableAccount.objects.filter(receivable_date__gte=period_start, receivable_date__lte=period_end)
        statistics['receivables'] = {
            'total': receivables.count(),
            'total_amount': receivables.aggregate(total=Sum('receivable_amount'))['total'] or Decimal('0'),
            'paid_amount': receivables.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0'),
            'remaining_amount': receivables.aggregate(total=Sum('remaining_amount'))['total'] or Decimal('0'),
        }
        
        # 应付账款统计
        payables = PayableAccount.objects.filter(payable_date__gte=period_start, payable_date__lte=period_end)
        statistics['payables'] = {
            'total': payables.count(),
            'total_amount': payables.aggregate(total=Sum('payable_amount'))['total'] or Decimal('0'),
            'paid_amount': payables.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0'),
            'remaining_amount': payables.aggregate(total=Sum('remaining_amount'))['total'] or Decimal('0'),
        }
        
        # 预算统计
        budgets = Budget.objects.filter(
            start_date__lte=period_end,
            end_date__gte=period_start
        )
        statistics['budgets'] = {
            'total': budgets.count(),
            'total_amount': budgets.aggregate(total=Sum('budget_amount'))['total'] or Decimal('0'),
            'used_amount': budgets.aggregate(total=Sum('used_amount'))['total'] or Decimal('0'),
            'remaining_amount': budgets.aggregate(total=Sum('remaining_amount'))['total'] or Decimal('0'),
        }
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取财务统计失败: %s', str(e))
        statistics = {}
    
    context = _context(
        "财务统计",
        "📊",
        f"查看 {period_start.strftime('%Y-%m-%d')} 至 {period_end.strftime('%Y-%m-%d')} 期间的财务统计数据",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'statistics': statistics,
        'period': period,
        'year': year,
        'month': month,
        'period_start': period_start,
        'period_end': period_end,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
        'quarters': range(1, 5),
    })
    return render(request, "financial_management/statistics.html", context)


@login_required
def account_subject_management(request):
    """会计科目管理"""
    permission_codes = get_user_permission_codes(request.user)
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    subject_type = request.GET.get('subject_type', '')
    is_active = request.GET.get('is_active', '')
    
    # 获取科目列表
    try:
        subjects = AccountSubject.objects.select_related('parent', 'created_by').order_by('code')
        
        # 应用筛选条件
        if search:
            subjects = subjects.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        if subject_type:
            subjects = subjects.filter(subject_type=subject_type)
        if is_active == 'true':
            subjects = subjects.filter(is_active=True)
        elif is_active == 'false':
            subjects = subjects.filter(is_active=False)
        
        # 分页
        paginator = Paginator(subjects, 50)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取会计科目列表失败: %s', str(e))
        page_obj = None
    
    # 统计信息
    try:
        total_subjects = AccountSubject.objects.count()
        active_subjects = AccountSubject.objects.filter(is_active=True).count()
        subjects_by_type = AccountSubject.objects.filter(is_active=True).values('subject_type').annotate(count=Count('id'))
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    context = _context(
        "会计科目管理",
        "📊",
        "管理会计科目信息",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'subjects': page_obj.object_list if page_obj else [],
        'subject_type_choices': AccountSubject.TYPE_CHOICES,
        'current_search': search,
        'current_subject_type': subject_type,
        'current_is_active': is_active,
    })
    return render(request, "financial_management/account_subject_list.html", context)


@login_required
def account_subject_tree_export(request):
    """导出会计科目树形结构"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.account.view', permission_codes):
        messages.error(request, '您没有权限导出会计科目')
        return redirect('finance_pages:account_subject_management')
    
    # 获取所有启用的一级科目
    root_subjects = AccountSubject.objects.filter(parent=None, is_active=True).order_by('code')
    
    # 创建Excel工作簿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '会计科目树'
    
    # 设置表头
    headers = ['科目编码', '科目名称', '科目类型', '余额方向', '级别', '是否启用', '备注说明']
    worksheet.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 递归导出科目树
    def export_subject_tree(subject, level=1):
        """递归导出科目及其子科目"""
        indent = '  ' * (level - 1)  # 根据级别添加缩进
        type_dict = dict(AccountSubject.TYPE_CHOICES)
        direction_dict = dict(AccountSubject.DIRECTION_CHOICES)
        
        row = [
            subject.code,
            indent + subject.name,  # 添加缩进显示层级关系
            type_dict.get(subject.subject_type, subject.subject_type),
            direction_dict.get(subject.direction, subject.direction),
            level,
            '是' if subject.is_active else '否',
            subject.description or '',
        ]
        worksheet.append(row)
        
        # 递归导出子科目
        children = AccountSubject.objects.filter(parent=subject, is_active=True).order_by('code')
        for child in children:
            export_subject_tree(child, level + 1)
    
    # 导出所有一级科目及其子科目
    for root_subject in root_subjects:
        export_subject_tree(root_subject)
    
    # 调整列宽
    column_widths = [15, 40, 15, 12, 8, 10, 30]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    # 设置数据行样式（根据级别设置不同的缩进和颜色）
    from openpyxl.styles import PatternFill as DataFill
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        level = row[4].value  # 级别列
        if level:
            # 根据级别设置不同的背景色
            if level == 1:
                fill_color = "E8F4F8"
            elif level == 2:
                fill_color = "F0F8E8"
            else:
                fill_color = "FFF8E8"
            
            for cell in row:
                cell.fill = DataFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('会计科目树_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def account_subject_import_template(request):
    """下载会计科目导入模板"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.account.manage', permission_codes):
        messages.error(request, '您没有权限下载导入模板')
        return redirect('finance_pages:account_subject_management')
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '会计科目导入模板'
    
    # 设置表头
    headers = ['科目编码', '科目名称', '上级科目编码', '科目类型', '余额方向', '是否启用', '备注说明']
    worksheet.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加示例数据
    examples = [
        ['1001', '库存现金', '', 'asset', 'debit', '是', ''],
        ['1002', '银行存款', '', 'asset', 'debit', '是', ''],
        ['100201', '工商银行', '1002', 'asset', 'debit', '是', ''],
        ['2001', '短期借款', '', 'liability', 'credit', '是', ''],
    ]
    for example in examples:
        worksheet.append(example)
    
    # 添加说明行
    worksheet.append([])
    worksheet.append(['说明：'])
    worksheet.append(['1. 科目编码：必填，唯一标识，不能重复'])
    worksheet.append(['2. 科目名称：必填'])
    worksheet.append(['3. 上级科目编码：可选，如果填写，系统会自动关联上级科目'])
    worksheet.append(['4. 科目类型：必填，可选值：asset(资产)、liability(负债)、equity(所有者权益)、revenue(收入)、expense(费用)、cost(成本)'])
    worksheet.append(['5. 余额方向：必填，可选值：debit(借方)、credit(贷方)'])
    worksheet.append(['6. 是否启用：必填，可选值：是、否'])
    
    # 调整列宽
    column_widths = [15, 20, 15, 15, 12, 12, 30]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="会计科目导入模板.xlsx"'
    workbook.save(response)
    return response


@login_required
def account_subject_import(request):
    """导入会计科目"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.account.manage', permission_codes):
        messages.error(request, '您没有权限导入会计科目')
        return redirect('finance_pages:account_subject_management')
    
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, '请选择要导入的文件')
            return redirect('finance_pages:account_subject_management')
        
        upload_file = request.FILES['file']
        if not upload_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, '请上传Excel文件（.xls或.xlsx格式）')
            return redirect('finance_pages:account_subject_management')
        
        try:
            from django.db import transaction
            workbook = load_workbook(upload_file, data_only=True)
            worksheet = workbook.active
            
            # 读取表头
            headers = [cell.value for cell in worksheet[1]]
            header_map = {str(h).strip(): i for i, h in enumerate(headers) if h}
            
            # 检查必填列
            required_columns = ['科目编码', '科目名称', '科目类型', '余额方向']
            missing_columns = [col for col in required_columns if col not in header_map]
            if missing_columns:
                messages.error(request, f'缺少必填列：{", ".join(missing_columns)}')
                return redirect('finance_pages:account_subject_management')
            
            success_count = 0
            error_count = 0
            errors = []
            
            # 读取数据行
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                # 跳过空行
                if not row or not any(row):
                    continue
                
                try:
                    code = str(row[header_map['科目编码']]).strip() if row[header_map['科目编码']] else None
                    name = str(row[header_map['科目名称']]).strip() if row[header_map['科目名称']] else None
                    
                    if not code or not name:
                        error_count += 1
                        errors.append(f'第{row_idx}行：科目编码和科目名称不能为空')
                        continue
                    
                    # 检查是否已存在
                    if AccountSubject.objects.filter(code=code).exists():
                        error_count += 1
                        errors.append(f'第{row_idx}行：科目编码 {code} 已存在')
                        continue
                    
                    # 解析其他字段
                    parent_code = str(row[header_map.get('上级科目编码', -1)]).strip() if header_map.get('上级科目编码', -1) >= 0 and row[header_map.get('上级科目编码', -1)] else None
                    subject_type = str(row[header_map['科目类型']]).strip() if row[header_map['科目类型']] else None
                    direction = str(row[header_map['余额方向']]).strip() if row[header_map['余额方向']] else None
                    is_active_str = str(row[header_map.get('是否启用', -1)]).strip() if header_map.get('是否启用', -1) >= 0 and row[header_map.get('是否启用', -1)] else '是'
                    description = str(row[header_map.get('备注说明', -1)]).strip() if header_map.get('备注说明', -1) >= 0 and row[header_map.get('备注说明', -1)] else ''
                    
                    # 验证科目类型
                    type_map = {
                        'asset': 'asset', '资产': 'asset',
                        'liability': 'liability', '负债': 'liability',
                        'equity': 'equity', '所有者权益': 'equity',
                        'revenue': 'revenue', '收入': 'revenue',
                        'expense': 'expense', '费用': 'expense',
                        'cost': 'cost', '成本': 'cost',
                    }
                    subject_type = type_map.get(subject_type.lower(), subject_type)
                    if subject_type not in dict(AccountSubject.TYPE_CHOICES):
                        error_count += 1
                        errors.append(f'第{row_idx}行：科目类型无效')
                        continue
                    
                    # 验证余额方向
                    direction_map = {
                        'debit': 'debit', '借方': 'debit',
                        'credit': 'credit', '贷方': 'credit',
                    }
                    direction = direction_map.get(direction.lower(), direction)
                    if direction not in dict(AccountSubject.DIRECTION_CHOICES):
                        error_count += 1
                        errors.append(f'第{row_idx}行：余额方向无效')
                        continue
                    
                    # 解析是否启用
                    is_active = is_active_str.lower() in ['是', 'yes', 'true', '1', 'y']
                    
                    # 查找上级科目
                    parent = None
                    level = 1
                    if parent_code:
                        try:
                            parent = AccountSubject.objects.get(code=parent_code)
                            level = parent.level + 1
                        except AccountSubject.DoesNotExist:
                            error_count += 1
                            errors.append(f'第{row_idx}行：上级科目编码 {parent_code} 不存在')
                            continue
                    
                    # 创建科目
                    with transaction.atomic():
                        AccountSubject.objects.create(
                            code=code,
                            name=name,
                            parent=parent,
                            subject_type=subject_type,
                            direction=direction,
                            level=level,
                            is_active=is_active,
                            description=description,
                            created_by=request.user,
                        )
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'第{row_idx}行：{str(e)}')
            
            if success_count > 0:
                messages.success(request, f'成功导入 {success_count} 个会计科目')
            if error_count > 0:
                error_msg = f'导入失败 {error_count} 个科目'
                if len(errors) <= 10:
                    error_msg += '：' + '；'.join(errors)
                else:
                    error_msg += f'：前10个错误：' + '；'.join(errors[:10])
                messages.warning(request, error_msg)
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('导入会计科目失败: %s', str(e))
            messages.error(request, f'导入失败：{str(e)}')
    
    return redirect('finance_pages:account_subject_management')


@login_required
def account_subject_create(request):
    """新增会计科目"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.account.manage', permission_codes):
        messages.error(request, '您没有权限新增会计科目')
        return redirect('finance_pages:account_subject_management')
    
    if request.method == 'POST':
        form = AccountSubjectForm(request.POST)
        if form.is_valid():
            account_subject = form.save(commit=False)
            account_subject.created_by = request.user
            # 如果选择了上级科目，自动计算级别
            if account_subject.parent:
                account_subject.level = account_subject.parent.level + 1
            account_subject.save()
            messages.success(request, f'会计科目 {account_subject.name} 创建成功！')
            return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject.id)
    else:
        form = AccountSubjectForm()
    
    context = _context(
        "新增会计科目",
        "➕",
        "创建新的会计科目",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/account_subject_form.html", context)


@login_required
def account_subject_update(request, account_subject_id):
    """编辑会计科目"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.account.manage', permission_codes):
        messages.error(request, '您没有权限编辑会计科目')
        return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject_id)
    
    account_subject = get_object_or_404(AccountSubject, id=account_subject_id)
    
    if request.method == 'POST':
        form = AccountSubjectForm(request.POST, instance=account_subject)
        if form.is_valid():
            account_subject = form.save(commit=False)
            # 如果选择了上级科目，自动计算级别
            if account_subject.parent:
                account_subject.level = account_subject.parent.level + 1
            account_subject.save()
            messages.success(request, f'会计科目 {account_subject.name} 更新成功！')
            return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject.id)
    else:
        form = AccountSubjectForm(instance=account_subject)
    
    context = _context(
        f"编辑会计科目 - {account_subject.name}",
        "✏️",
        f"编辑会计科目 {account_subject.name}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'account_subject': account_subject,
        'is_create': False,
    })
    return render(request, "financial_management/account_subject_form.html", context)


@login_required
def account_subject_delete(request, account_subject_id):
    """删除会计科目"""
    account_subject = get_object_or_404(AccountSubject, id=account_subject_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.account.manage', permission_codes):
        messages.error(request, '您没有权限删除会计科目')
        return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject_id)
    
    # 检查是否有子科目
    children_count = account_subject.children.count()
    if children_count > 0:
        messages.error(request, f'该科目下有 {children_count} 个子科目，无法删除。请先删除或移动子科目。')
        return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject_id)
    
    # 检查是否被使用
    voucher_entry_count = account_subject.voucher_entries.count()
    ledger_entry_count = account_subject.ledger_entries.count()
    
    if voucher_entry_count > 0 or ledger_entry_count > 0:
        messages.error(request, f'该科目已被使用（凭证分录：{voucher_entry_count}条，总账记录：{ledger_entry_count}条），无法删除。')
        return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject_id)
    
    if request.method == 'POST':
        try:
            subject_name = account_subject.name
            account_subject.delete()
            messages.success(request, f'会计科目 {subject_name} 已删除')
            return redirect('finance_pages:account_subject_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除会计科目失败: %s', str(e))
            messages.error(request, f'删除会计科目失败：{str(e)}')
            return redirect('finance_pages:account_subject_detail', account_subject_id=account_subject_id)
    
    context = _context(
        f"删除会计科目 - {account_subject.name}",
        "🗑️",
        f"确认删除会计科目：{account_subject.code} {account_subject.name}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'account_subject': account_subject,
    })
    return render(request, "financial_management/account_subject_delete.html", context)


@login_required
def budget_create(request):
    """新增预算"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.budget.create', permission_codes):
        messages.error(request, '您没有权限创建预算')
        return redirect('finance_pages:budget_management')
    
    if request.method == 'POST':
        form = BudgetForm(request.POST)
        if form.is_valid():
            budget = form.save(commit=False)
            # 自动生成预算编号
            if not budget.budget_number:
                current_year = timezone.now().year
                max_budget = Budget.objects.filter(
                    budget_number__startswith=f'BUDGET-{current_year}-'
                ).aggregate(max_num=Max('budget_number'))['max_num']
                if max_budget:
                    try:
                        seq = int(max_budget.split('-')[-1]) + 1
                    except (ValueError, IndexError):
                        seq = 1
                else:
                    seq = 1
                budget.budget_number = f'BUDGET-{current_year}-{seq:04d}'
            budget.remaining_amount = budget.budget_amount
            budget.created_by = request.user
            budget.save()
            messages.success(request, f'预算 {budget.name} 创建成功！')
            return redirect('finance_pages:budget_detail', budget_id=budget.id)
    else:
        form = BudgetForm()
    
    context = _context(
        "新增预算",
        "➕",
        "创建新的预算记录",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/budget_form.html", context)


@login_required
def budget_update(request, budget_id):
    """编辑预算"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.budget.manage', permission_codes):
        messages.error(request, '您没有权限编辑预算')
        return redirect('finance_pages:budget_detail', budget_id=budget_id)
    
    budget = get_object_or_404(Budget, id=budget_id)
    
    if request.method == 'POST':
        form = BudgetForm(request.POST, instance=budget)
        if form.is_valid():
            budget = form.save(commit=False)
            # 重新计算剩余金额
            budget.remaining_amount = budget.budget_amount - budget.used_amount
            budget.save()
            messages.success(request, f'预算 {budget.name} 更新成功！')
            return redirect('finance_pages:budget_detail', budget_id=budget.id)
    else:
        form = BudgetForm(instance=budget)
    
    context = _context(
        f"编辑预算 - {budget.name}",
        "✏️",
        f"编辑预算 {budget.name}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'budget': budget,
        'is_create': False,
    })
    return render(request, "financial_management/budget_form.html", context)


@login_required
def invoice_create(request):
    """新增发票"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.invoice.create', permission_codes):
        messages.error(request, '您没有权限创建发票')
        return redirect('finance_pages:invoice_management')
    
    if request.method == 'POST':
        form = InvoiceForm(request.POST, request.FILES)
        if form.is_valid():
            invoice = form.save(commit=False)
            # 如果没有填写总金额，自动计算
            if not invoice.total_amount and invoice.amount and invoice.tax_amount:
                invoice.total_amount = invoice.amount + invoice.tax_amount
            elif not invoice.total_amount:
                invoice.total_amount = invoice.amount or Decimal('0.00')
            invoice.created_by = request.user
            invoice.save()
            messages.success(request, f'发票 {invoice.invoice_number} 创建成功！')
            return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    else:
        form = InvoiceForm()
        # 默认当前日期
        form.fields['invoice_date'].initial = timezone.now().date()
    
    context = _context(
        "新增发票",
        "➕",
        "创建新的发票记录",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/invoice_form.html", context)


@login_required
def invoice_update(request, invoice_id):
    """编辑发票"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.invoice.manage', permission_codes):
        messages.error(request, '您没有权限编辑发票')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice_id)
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if request.method == 'POST':
        form = InvoiceForm(request.POST, request.FILES, instance=invoice)
        if form.is_valid():
            invoice = form.save(commit=False)
            # 重新计算总金额
            if invoice.amount and invoice.tax_amount:
                invoice.total_amount = invoice.amount + invoice.tax_amount
            invoice.save()
            messages.success(request, f'发票 {invoice.invoice_number} 更新成功！')
            return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    else:
        form = InvoiceForm(instance=invoice)
    
    context = _context(
        f"编辑发票 - {invoice.invoice_number}",
        "✏️",
        f"编辑发票 {invoice.invoice_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'invoice': invoice,
        'is_create': False,
    })
    return render(request, "financial_management/invoice_form.html", context)


@login_required
def fund_flow_create(request):
    """新增资金流水"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.fund_flow.create', permission_codes):
        messages.error(request, '您没有权限创建资金流水')
        return redirect('finance_pages:fund_flow_management')
    
    if request.method == 'POST':
        form = FundFlowForm(request.POST)
        if form.is_valid():
            fund_flow = form.save(commit=False)
            # 自动生成流水号
            if not fund_flow.flow_number:
                current_year = timezone.now().year
                max_flow = FundFlow.objects.filter(
                    flow_number__startswith=f'FLOW-{current_year}-'
                ).aggregate(max_num=Max('flow_number'))['max_num']
                if max_flow:
                    try:
                        seq = int(max_flow.split('-')[-1]) + 1
                    except (ValueError, IndexError):
                        seq = 1
                else:
                    seq = 1
                fund_flow.flow_number = f'FLOW-{current_year}-{seq:04d}'
            fund_flow.created_by = request.user
            fund_flow.save()
            
            # 自动更新预算使用金额
            _update_budget_from_fund_flow(fund_flow, is_create=True)
            
            messages.success(request, f'资金流水 {fund_flow.flow_number} 创建成功！')
            return redirect('finance_pages:fund_flow_detail', fund_flow_id=fund_flow.id)
    else:
        form = FundFlowForm()
        # 默认今天
        form.fields['flow_date'].initial = timezone.now().date()
    
    context = _context(
        "新增资金流水",
        "➕",
        "创建新的资金流水记录",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/fund_flow_form.html", context)


@login_required
def fund_flow_update(request, fund_flow_id):
    """编辑资金流水"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.fund_flow.manage', permission_codes):
        messages.error(request, '您没有权限编辑资金流水')
        return redirect('finance_pages:fund_flow_detail', fund_flow_id=fund_flow_id)
    
    fund_flow = get_object_or_404(FundFlow, id=fund_flow_id)
    
    if request.method == 'POST':
        form = FundFlowForm(request.POST, instance=fund_flow)
        if form.is_valid():
            # 保存旧金额用于预算更新
            old_amount = fund_flow.amount
            old_flow_type = fund_flow.flow_type
            
            fund_flow = form.save()
            
            # 如果金额或类型发生变化，更新预算
            if old_amount != fund_flow.amount or old_flow_type != fund_flow.flow_type:
                # 先回滚旧金额的影响
                if old_flow_type == 'expense':
                    _update_budget_from_fund_flow(fund_flow, is_create=False, old_amount=old_amount)
                # 再应用新金额的影响
                _update_budget_from_fund_flow(fund_flow, is_create=True)
            
            messages.success(request, f'资金流水 {fund_flow.flow_number} 更新成功！')
            return redirect('finance_pages:fund_flow_detail', fund_flow_id=fund_flow.id)
    else:
        form = FundFlowForm(instance=fund_flow)
    
    context = _context(
        f"编辑资金流水 - {fund_flow.flow_number}",
        "✏️",
        f"编辑资金流水 {fund_flow.flow_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'fund_flow': fund_flow,
        'is_create': False,
    })
    return render(request, "financial_management/fund_flow_form.html", context)


# 创建凭证分录的内联表单集
VoucherEntryFormSet = inlineformset_factory(
    Voucher, VoucherEntry,
    form=VoucherEntryForm,
    extra=3,  # 默认显示3个空行
    can_delete=True,
    min_num=1,  # 至少需要1个分录
    validate_min=True,
)


@login_required
def voucher_create(request):
    """新增记账凭证"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.create', permission_codes):
        messages.error(request, '您没有权限创建记账凭证')
        return redirect('finance_pages:voucher_management')
    
    if request.method == 'POST':
        form = VoucherForm(request.POST)
        formset = VoucherEntryFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            voucher = form.save(commit=False)
            # 自动生成凭证字号
            if not voucher.voucher_number:
                voucher.voucher_number = _generate_voucher_number(voucher.voucher_date)
            if not voucher.preparer:
                voucher.preparer = request.user
            voucher.save()
            
            # 保存分录并计算合计
            entries = formset.save(commit=False)
            total_debit = Decimal('0.00')
            total_credit = Decimal('0.00')
            
            for entry in entries:
                entry.voucher = voucher
                entry.save()
                total_debit += entry.debit_amount or Decimal('0.00')
                total_credit += entry.credit_amount or Decimal('0.00')
            
            # 删除标记为删除的分录
            for obj in formset.deleted_objects:
                obj.delete()
            
            # 更新合计
            voucher.total_debit = total_debit
            voucher.total_credit = total_credit
            voucher.save()
            
            messages.success(request, f'记账凭证 {voucher.voucher_number} 创建成功！')
            return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
        else:
            messages.error(request, '请检查表单中的错误。')
    else:
        form = VoucherForm(initial={'voucher_date': timezone.now().date(), 'preparer': request.user})
        formset = VoucherEntryFormSet()
    
    # 获取所有会计科目供 JavaScript 使用
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    
    context = _context(
        "新增记账凭证",
        "➕",
        "创建新的记账凭证",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'formset': formset,
        'is_create': True,
        'account_subjects': account_subjects,
    })
    return render(request, "financial_management/voucher_form.html", context)


@login_required
def voucher_update(request, voucher_id):
    """编辑记账凭证"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        messages.error(request, '您没有权限编辑记账凭证')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher.objects.prefetch_related('entries'), id=voucher_id)
    
    # 已过账的凭证不能编辑
    if voucher.status == 'posted':
        messages.error(request, '已过账的凭证不能编辑')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        form = VoucherForm(request.POST, instance=voucher)
        formset = VoucherEntryFormSet(request.POST, instance=voucher)
        
        if form.is_valid() and formset.is_valid():
            voucher = form.save()
            
            # 保存分录并计算合计
            entries = formset.save(commit=False)
            total_debit = Decimal('0.00')
            total_credit = Decimal('0.00')
            
            for entry in entries:
                entry.voucher = voucher
                entry.save()
                total_debit += entry.debit_amount or Decimal('0.00')
                total_credit += entry.credit_amount or Decimal('0.00')
            
            # 删除标记为删除的分录
            for obj in formset.deleted_objects:
                obj.delete()
            
            # 更新合计
            voucher.total_debit = total_debit
            voucher.total_credit = total_credit
            voucher.save()
            
            messages.success(request, f'记账凭证 {voucher.voucher_number} 更新成功！')
            return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
        else:
            messages.error(request, '请检查表单中的错误。')
    else:
        form = VoucherForm(instance=voucher)
        formset = VoucherEntryFormSet(instance=voucher)
    
    # 获取所有会计科目供 JavaScript 使用
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    
    context = _context(
        f"编辑记账凭证 - {voucher.voucher_number}",
        "✏️",
        f"编辑记账凭证 {voucher.voucher_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'formset': formset,
        'voucher': voucher,
        'is_create': False,
        'account_subjects': account_subjects,
    })
    return render(request, "financial_management/voucher_form.html", context)


@login_required
def voucher_management(request):
    """凭证管理"""
    permission_codes = get_user_permission_codes(request.user)
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # 获取凭证列表（用于统计，在筛选之前）
    base_vouchers = Voucher.objects.select_related('preparer', 'reviewer', 'posted_by').all()
    
    # 统计数据（在过滤之前获取，显示全部数据统计）
    total_count = base_vouchers.count()
    draft_count = base_vouchers.filter(status='draft').count()
    submitted_count = base_vouchers.filter(status='submitted').count()
    approved_count = base_vouchers.filter(status='approved').count()
    posted_count = base_vouchers.filter(status='posted').count()
    rejected_count = base_vouchers.filter(status='rejected').count()
    
    # 应用筛选条件
    vouchers = base_vouchers.order_by('-voucher_date', '-voucher_number')
    
    if search:
        vouchers = vouchers.filter(
            Q(voucher_number__icontains=search) |
            Q(notes__icontains=search)
        )
    if status:
        vouchers = vouchers.filter(status=status)
    if date_from:
        vouchers = vouchers.filter(voucher_date__gte=date_from)
    if date_to:
        vouchers = vouchers.filter(voucher_date__lte=date_to)
    
    # 分页（每页20条）
    paginator = Paginator(vouchers, 20)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = paginator.get_page(1)
    
    # 生成左侧菜单
    financial_menu = _build_financial_sidebar_nav(permission_codes, request.path)
    
    context = _context(
        "凭证管理",
        "📝",
        "管理记账凭证",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'page': page_obj,  # 兼容模板中的变量名
        'vouchers': page_obj.object_list if page_obj else [],
        'status_choices': Voucher.STATUS_CHOICES,
        'status_filter': status,  # 兼容模板中的变量名
        'search': search,
        'current_search': search,  # 兼容模板中的变量名
        'current_status': status,  # 兼容模板中的变量名
        'status': status,  # 兼容模板中的变量名
        'date_from': date_from,
        'date_to': date_to,
        'current_date_from': date_from,  # 兼容模板中的变量名
        'current_date_to': date_to,  # 兼容模板中的变量名
        'total_count': total_count,
        'draft_count': draft_count,
        'submitted_count': submitted_count,
        'approved_count': approved_count,
        'posted_count': posted_count,
        'rejected_count': rejected_count,
        'financial_menu': financial_menu,
        'module_sidebar_nav': financial_menu,  # 兼容模板中的变量名
        'sidebar_title': '财务管理',  # 侧边栏标题
        'sidebar_subtitle': 'Financial Management',  # 侧边栏副标题
    })
    return render(request, "financial_management/voucher_list.html", context)


@login_required
def voucher_export(request):
    """导出凭证列表为Excel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.view', permission_codes):
        messages.error(request, '您没有权限导出凭证')
        return redirect('finance_pages:voucher_management')
    
    # 获取筛选参数（与列表页保持一致）
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # 获取凭证列表
    vouchers = Voucher.objects.select_related('preparer', 'reviewer', 'posted_by').order_by('-voucher_date', '-voucher_number')
    
    if search:
        vouchers = vouchers.filter(
            Q(voucher_number__icontains=search) |
            Q(notes__icontains=search)
        )
    if status:
        vouchers = vouchers.filter(status=status)
    if date_from:
        vouchers = vouchers.filter(voucher_date__gte=date_from)
    if date_to:
        vouchers = vouchers.filter(voucher_date__lte=date_to)
    
    # 创建Excel工作簿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '凭证列表'
    
    # 设置表头
    headers = ['凭证字号', '凭证日期', '借方合计', '贷方合计', '状态', '制单人', '审核人', '审核时间', '过账人', '过账时间', '附件数', '备注']
    worksheet.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加数据
    status_dict = dict(Voucher.STATUS_CHOICES)
    for voucher in vouchers:
        row = [
            voucher.voucher_number,
            voucher.voucher_date.strftime('%Y-%m-%d') if voucher.voucher_date else '',
            float(voucher.total_debit),
            float(voucher.total_credit),
            status_dict.get(voucher.status, voucher.status),
            voucher.preparer.get_full_name() if voucher.preparer else '',
            voucher.reviewer.get_full_name() if voucher.reviewer else '',
            voucher.reviewed_time.strftime('%Y-%m-%d %H:%M') if voucher.reviewed_time else '',
            voucher.posted_by.get_full_name() if voucher.posted_by else '',
            voucher.posted_time.strftime('%Y-%m-%d %H:%M') if voucher.posted_time else '',
            voucher.attachment_count,
            voucher.notes or '',
        ]
        worksheet.append(row)
    
    # 调整列宽
    column_widths = [18, 12, 12, 12, 10, 12, 12, 18, 12, 18, 10, 30]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    # 生成响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('凭证列表_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def voucher_batch_approve(request):
    """批量审核凭证"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.review', permission_codes):
        messages.error(request, '您没有权限批量审核凭证')
        return redirect('finance_pages:voucher_management')
    
    if request.method == 'POST':
        voucher_ids = request.POST.getlist('voucher_ids')
        action = request.POST.get('action', 'approve')
        
        if not voucher_ids:
            messages.error(request, '请选择要操作的凭证')
            return redirect('finance_pages:voucher_management')
        
        try:
            vouchers = Voucher.objects.filter(id__in=voucher_ids)
            success_count = 0
            error_count = 0
            
            for voucher in vouchers:
                # 检查状态
                if voucher.status not in ['submitted', 'draft']:
                    error_count += 1
                    continue
                
                # 检查借贷是否平衡
                if voucher.total_debit != voucher.total_credit:
                    error_count += 1
                    continue
                
                if action == 'approve':
                    voucher.status = 'approved'
                    voucher.reviewer = request.user
                    voucher.reviewed_time = timezone.now()
                    voucher.save()
                    success_count += 1
                elif action == 'reject':
                    voucher.status = 'rejected'
                    voucher.reviewer = request.user
                    voucher.reviewed_time = timezone.now()
                    voucher.save()
                    success_count += 1
            
            if success_count > 0:
                messages.success(request, f'成功{action == "approve" and "审核通过" or "拒绝"} {success_count} 张凭证')
            if error_count > 0:
                messages.warning(request, f'{error_count} 张凭证操作失败（状态不符合要求或借贷不平衡）')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('批量审核凭证失败: %s', str(e))
            messages.error(request, f'批量操作失败：{str(e)}')
    
    return redirect('finance_pages:voucher_management')


@login_required
def voucher_batch_post(request):
    """批量过账凭证"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.post', permission_codes):
        messages.error(request, '您没有权限批量过账凭证')
        return redirect('finance_pages:voucher_management')
    
    if request.method == 'POST':
        voucher_ids = request.POST.getlist('voucher_ids')
        
        if not voucher_ids:
            messages.error(request, '请选择要过账的凭证')
            return redirect('finance_pages:voucher_management')
        
        try:
            vouchers = Voucher.objects.filter(id__in=voucher_ids)
            success_count = 0
            error_count = 0
            
            for voucher in vouchers:
                # 检查状态
                if voucher.status != 'approved':
                    error_count += 1
                    continue
                
                # 检查借贷是否平衡
                if voucher.total_debit != voucher.total_credit:
                    error_count += 1
                    continue
                
                # 执行过账
                try:
                    from django.db import transaction
                    with transaction.atomic():
                        # 更新凭证状态
                        voucher.status = 'posted'
                        voucher.posted_by = request.user
                        voucher.posted_time = timezone.now()
                        voucher.save()
                        
                        # 生成总账记录
                        entries = voucher.entries.select_related('account_subject').all()
                        period_year = voucher.voucher_date.year
                        period_month = voucher.voucher_date.month
                        
                        for entry in entries:
                            # 获取或创建总账记录
                            ledger, created = Ledger.objects.get_or_create(
                                account_subject=entry.account_subject,
                                period_year=period_year,
                                period_month=period_month,
                                period_date=voucher.voucher_date,
                                defaults={
                                    'opening_balance': Decimal('0.00'),
                                    'period_debit': Decimal('0.00'),
                                    'period_credit': Decimal('0.00'),
                                    'closing_balance': Decimal('0.00'),
                                }
                            )
                            
                            # 更新总账
                            ledger.period_debit += entry.debit_amount or Decimal('0.00')
                            ledger.period_credit += entry.credit_amount or Decimal('0.00')
                            
                            # 计算期末余额
                            if entry.account_subject.direction == 'debit':
                                ledger.closing_balance = ledger.opening_balance + ledger.period_debit - ledger.period_credit
                            else:
                                ledger.closing_balance = ledger.opening_balance + ledger.period_credit - ledger.period_debit
                            
                            ledger.save()
                        
                        success_count += 1
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.exception('过账凭证失败: %s', str(e))
                    error_count += 1
            
            if success_count > 0:
                messages.success(request, f'成功过账 {success_count} 张凭证')
            if error_count > 0:
                messages.warning(request, f'{error_count} 张凭证过账失败')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('批量过账凭证失败: %s', str(e))
            messages.error(request, f'批量过账失败：{str(e)}')
    
    return redirect('finance_pages:voucher_management')


@login_required
def ledger_management(request):
    """账簿管理"""
    permission_codes = get_user_permission_codes(request.user)
    today = timezone.now().date()
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', '')
    account_subject_id = request.GET.get('account_subject_id', '')
    
    # 获取总账列表
    try:
        ledgers = Ledger.objects.select_related('account_subject').order_by('-period_date', 'account_subject__code')
        
        # 应用筛选条件
        if search:
            ledgers = ledgers.filter(
                Q(account_subject__code__icontains=search) |
                Q(account_subject__name__icontains=search)
            )
        if period_year:
            ledgers = ledgers.filter(period_year=int(period_year))
        if period_month:
            ledgers = ledgers.filter(period_month=int(period_month))
        if account_subject_id:
            ledgers = ledgers.filter(account_subject_id=int(account_subject_id))
        
        # 分页
        paginator = Paginator(ledgers, 50)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取总账列表失败: %s', str(e))
        page_obj = None
    
    # 统计信息
    try:
        current_year = int(period_year) if period_year else today.year
        current_month = int(period_month) if period_month else today.month
        ledger_count = Ledger.objects.filter(
            period_year=current_year,
            period_month=current_month
        ).count()
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    context = _context(
        "账簿管理",
        "📖",
        "管理总账、明细账等",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'ledgers': page_obj.object_list if page_obj else [],
        'current_search': search,
        'current_period_year': period_year,
        'current_period_month': period_month,
        'current_account_subject_id': account_subject_id,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/ledger_list.html", context)


@login_required
def budget_management(request):
    """预算管理"""
    permission_codes = get_user_permission_codes(request.user)
    today = timezone.now().date()
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    budget_year = request.GET.get('budget_year', '')
    
    # 获取预算列表
    try:
        budgets = Budget.objects.select_related('department', 'account_subject', 'approver', 'created_by').order_by('-budget_year', '-created_time')
        
        # 应用筛选条件
        if search:
            budgets = budgets.filter(
                Q(budget_number__icontains=search) |
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        if status:
            budgets = budgets.filter(status=status)
        if budget_year:
            budgets = budgets.filter(budget_year=int(budget_year))
        
        # 分页
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(budgets, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取预算列表失败: %s', str(e))
        page_obj = None
    
    # 统计信息
    try:
        total_budgets = Budget.objects.count()
        executing_budgets = Budget.objects.filter(status='executing').count()
        total_budget_amount = Budget.objects.filter(status='executing').aggregate(
            total=Sum('budget_amount')
        )['total'] or Decimal('0')
        total_used_amount = Budget.objects.filter(status='executing').aggregate(
            total=Sum('used_amount')
        )['total'] or Decimal('0')
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    context = _context(
        "预算管理",
        "💰",
        "管理预算编制和执行",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'budgets': page_obj.object_list if page_obj else [],
        'status_choices': Budget.STATUS_CHOICES,
        'current_search': search,
        'current_status': status,
        'current_budget_year': budget_year,
        'years': range(today.year - 2, today.year + 2),
    })
    return render(request, "financial_management/budget_list.html", context)


@login_required
def budget_execution_analysis(request):
    """预算执行情况分析"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.budget.view', permission_codes):
        messages.error(request, '您没有权限查看预算执行分析')
        return redirect('finance_pages:budget_management')
    
    today = timezone.now().date()
    
    # 获取筛选参数
    budget_year = int(request.GET.get('budget_year', today.year))
    department_id = request.GET.get('department_id', '')
    account_subject_id = request.GET.get('account_subject_id', '')
    
    # 获取预算数据
    budgets = Budget.objects.select_related('department', 'account_subject', 'approver', 'created_by').filter(
        budget_year=budget_year,
        status__in=['approved', 'executing']
    )
    
    if department_id:
        budgets = budgets.filter(department_id=department_id)
    if account_subject_id:
        budgets = budgets.filter(account_subject_id=account_subject_id)
    
    # 计算执行情况
    execution_data = []
    total_budget = Decimal('0')
    total_used = Decimal('0')
    total_remaining = Decimal('0')
    
    for budget in budgets:
        usage_rate = 0
        if budget.budget_amount > 0:
            usage_rate = (budget.used_amount / budget.budget_amount) * 100
        
        remaining_rate = 0
        if budget.budget_amount > 0:
            remaining_rate = (budget.remaining_amount / budget.budget_amount) * 100
        
        execution_data.append({
            'budget': budget,
            'usage_rate': usage_rate,
            'remaining_rate': remaining_rate,
            'is_over_budget': budget.used_amount > budget.budget_amount,
            'is_near_limit': usage_rate >= 80 and usage_rate < 100,
        })
        
        total_budget += budget.budget_amount
        total_used += budget.used_amount
        total_remaining += budget.remaining_amount
    
    # 计算总体执行率
    overall_usage_rate = 0
    if total_budget > 0:
        overall_usage_rate = (total_used / total_budget) * 100
    
    # 按执行率排序
    execution_data.sort(key=lambda x: x['usage_rate'], reverse=True)
    
    # 统计信息
    summary_cards = []
    
    # 获取部门列表和科目列表供筛选
    try:
        from backend.apps.system_management.models import Department
        departments = Department.objects.filter(is_active=True).order_by('name')
    except Exception:
        departments = []
    
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    
    context = _context(
        f"预算执行情况分析 - {budget_year}年",
        "📊",
        f"分析 {budget_year} 年度预算执行情况",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'execution_data': execution_data,
        'budget_year': budget_year,
        'total_budget': total_budget,
        'total_used': total_used,
        'total_remaining': total_remaining,
        'overall_usage_rate': overall_usage_rate,
        'departments': departments,
        'account_subjects': account_subjects,
        'current_department_id': department_id,
        'current_account_subject_id': account_subject_id,
        'years': range(today.year - 2, today.year + 2),
    })
    return render(request, "financial_management/budget_execution_analysis.html", context)


@login_required
def budget_export(request):
    """导出预算列表为Excel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.budget.view', permission_codes):
        messages.error(request, '您没有权限导出预算')
        return redirect('finance_pages:budget_management')
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    budget_year = request.GET.get('budget_year', '')
    
    budgets = Budget.objects.select_related('department', 'account_subject', 'approver', 'created_by').order_by('-budget_year', '-created_time')
    
    if search:
        budgets = budgets.filter(
            Q(budget_number__icontains=search) |
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    if status:
        budgets = budgets.filter(status=status)
    if budget_year:
        budgets = budgets.filter(budget_year=int(budget_year))
    
    # 创建Excel工作簿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '预算列表'
    
    headers = ['预算编号', '预算名称', '预算年度', '预算金额', '已用金额', '剩余金额', '所属部门', '预算科目', '状态', '审批人', '审批时间', '开始日期', '结束日期', '创建人', '创建时间']
    worksheet.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加数据
    status_dict = dict(Budget.STATUS_CHOICES)
    for budget in budgets:
        row = [
            budget.budget_number,
            budget.name,
            budget.budget_year,
            float(budget.budget_amount),
            float(budget.used_amount),
            float(budget.remaining_amount),
            budget.department.name if budget.department else '',
            f"{budget.account_subject.code} {budget.account_subject.name}" if budget.account_subject else '',
            status_dict.get(budget.status, budget.status),
            budget.approver.get_full_name() if budget.approver else '',
            budget.approved_time.strftime('%Y-%m-%d %H:%M') if budget.approved_time else '',
            budget.start_date.strftime('%Y-%m-%d') if budget.start_date else '',
            budget.end_date.strftime('%Y-%m-%d') if budget.end_date else '',
            budget.created_by.get_full_name() if budget.created_by else '',
            budget.created_time.strftime('%Y-%m-%d %H:%M') if budget.created_time else '',
        ]
        worksheet.append(row)
    
    # 调整列宽
    column_widths = [18, 25, 10, 12, 12, 12, 15, 20, 10, 12, 18, 12, 12, 12, 18]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('预算列表_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def invoice_management(request):
    """发票管理"""
    permission_codes = get_user_permission_codes(request.user)
    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    invoice_type = request.GET.get('invoice_type', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # 获取发票列表
    try:
        invoices = Invoice.objects.select_related('verified_by', 'created_by').order_by('-invoice_date', '-invoice_number')
        
        # 应用筛选条件
        if search:
            invoices = invoices.filter(
                Q(invoice_number__icontains=search) |
                Q(invoice_code__icontains=search) |
                Q(customer_name__icontains=search) |
                Q(supplier_name__icontains=search)
            )
        if invoice_type:
            invoices = invoices.filter(invoice_type=invoice_type)
        if status:
            invoices = invoices.filter(status=status)
        if date_from:
            invoices = invoices.filter(invoice_date__gte=date_from)
        if date_to:
            invoices = invoices.filter(invoice_date__lte=date_to)
        
        # 分页
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(invoices, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取发票列表失败: %s', str(e))
        page_obj = None
    
    # 统计信息
    try:
        total_invoices = Invoice.objects.count()
        issued_invoices = Invoice.objects.filter(status='issued').count()
        this_month_invoices = Invoice.objects.filter(invoice_date__gte=this_month_start).count()
        this_month_amount = Invoice.objects.filter(invoice_date__gte=this_month_start).aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0')
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    context = _context(
        "发票管理",
        "🧾",
        "管理进项和销项发票",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'invoices': page_obj.object_list if page_obj else [],
        'invoice_type_choices': Invoice.TYPE_CHOICES,
        'status_choices': Invoice.STATUS_CHOICES,
        'current_search': search,
        'current_invoice_type': invoice_type,
        'current_status': status,
        'current_date_from': date_from,
        'current_date_to': date_to,
    })
    return render(request, "financial_management/invoice_list.html", context)


@login_required
def invoice_export(request):
    """导出发票列表为Excel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.invoice.view', permission_codes):
        messages.error(request, '您没有权限导出发票')
        return redirect('finance_pages:invoice_management')
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    invoice_type = request.GET.get('invoice_type', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    invoices = Invoice.objects.select_related('project', 'created_by').order_by('-invoice_date', '-invoice_number')
    
    if search:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search) |
            Q(seller_name__icontains=search) |
            Q(buyer_name__icontains=search) |
            Q(description__icontains=search)
        )
    if invoice_type:
        invoices = invoices.filter(invoice_type=invoice_type)
    if status:
        invoices = invoices.filter(status=status)
    if date_from:
        invoices = invoices.filter(invoice_date__gte=date_from)
    if date_to:
        invoices = invoices.filter(invoice_date__lte=date_to)
    
    # 创建Excel工作簿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '发票列表'
    
    headers = ['发票号码', '发票类型', '发票日期', '销售方', '购买方', '金额', '税额', '价税合计', '状态', '关联项目', '备注', '创建人', '创建时间']
    worksheet.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加数据
    type_dict = dict(Invoice.TYPE_CHOICES)
    status_dict = dict(Invoice.STATUS_CHOICES)
    for invoice in invoices:
        row = [
            invoice.invoice_number,
            type_dict.get(invoice.invoice_type, invoice.invoice_type),
            invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else '',
            invoice.seller_name or '',
            invoice.buyer_name or '',
            float(invoice.amount) if invoice.amount else 0,
            float(invoice.tax_amount) if invoice.tax_amount else 0,
            float(invoice.total_amount) if invoice.total_amount else 0,
            status_dict.get(invoice.status, invoice.status),
            invoice.project.project_number if invoice.project else '',
            invoice.description or '',
            invoice.created_by.get_full_name() if invoice.created_by else '',
            invoice.created_time.strftime('%Y-%m-%d %H:%M') if invoice.created_time else '',
        ]
        worksheet.append(row)
    
    # 调整列宽
    column_widths = [20, 10, 12, 20, 20, 12, 12, 12, 10, 15, 30, 12, 18]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('发票列表_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def fund_flow_management(request):
    """资金流水"""
    permission_codes = get_user_permission_codes(request.user)
    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    flow_type = request.GET.get('flow_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # 获取资金流水列表
    try:
        fund_flows = FundFlow.objects.select_related('project', 'voucher', 'created_by').order_by('-flow_date', '-flow_number')
        
        # 应用筛选条件
        if search:
            fund_flows = fund_flows.filter(
                Q(flow_number__icontains=search) |
                Q(account_name__icontains=search) |
                Q(counterparty__icontains=search) |
                Q(summary__icontains=search)
            )
        if flow_type:
            fund_flows = fund_flows.filter(flow_type=flow_type)
        if date_from:
            fund_flows = fund_flows.filter(flow_date__gte=date_from)
        if date_to:
            fund_flows = fund_flows.filter(flow_date__lte=date_to)
        
        # 分页
        paginator = Paginator(fund_flows, 50)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取资金流水列表失败: %s', str(e))
        page_obj = None
    
    # 统计信息
    try:
        this_month_flows = FundFlow.objects.filter(flow_date__gte=this_month_start).count()
        this_month_income = FundFlow.objects.filter(
            flow_date__gte=this_month_start,
            flow_type='income'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        this_month_expense = FundFlow.objects.filter(
            flow_date__gte=this_month_start,
            flow_type='expense'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    context = _context(
        "资金流水",
        "💳",
        "管理资金流入流出记录",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'fund_flows': page_obj.object_list if page_obj else [],
        'flow_type_choices': FundFlow.TYPE_CHOICES,
        'current_search': search,
        'current_flow_type': flow_type,
        'current_date_from': date_from,
        'current_date_to': date_to,
    })
    return render(request, "financial_management/fund_flow_list.html", context)


@login_required
def fund_flow_import_template(request):
    """下载资金流水导入模板"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.fund_flow.create', permission_codes):
        messages.error(request, '您没有权限下载导入模板')
        return redirect('finance_pages:fund_flow_management')
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '资金流水导入模板'
    
    # 设置表头
    headers = ['发生日期', '流水类型', '金额', '账户名称', '对方单位', '摘要', '关联项目编号']
    worksheet.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加示例数据
    examples = [
        ['2025-01-15', 'income', '10000.00', '工商银行', '客户A', '项目回款', 'PJT-2025-001'],
        ['2025-01-16', 'expense', '5000.00', '工商银行', '供应商B', '材料采购', 'PJT-2025-002'],
    ]
    for example in examples:
        worksheet.append(example)
    
    # 添加说明行
    worksheet.append([])
    worksheet.append(['说明：'])
    worksheet.append(['1. 发生日期：必填，格式：YYYY-MM-DD'])
    worksheet.append(['2. 流水类型：必填，可选值：income(收入)、expense(支出)、transfer(转账)'])
    worksheet.append(['3. 金额：必填，数字格式'])
    worksheet.append(['4. 账户名称：必填'])
    worksheet.append(['5. 对方单位：可选'])
    worksheet.append(['6. 摘要：必填'])
    worksheet.append(['7. 关联项目编号：可选，如果填写，系统会自动关联项目'])
    
    # 调整列宽
    column_widths = [15, 12, 15, 20, 20, 30, 20]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="资金流水导入模板.xlsx"'
    workbook.save(response)
    return response


@login_required
def fund_flow_import(request):
    """导入资金流水"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.fund_flow.create', permission_codes):
        messages.error(request, '您没有权限导入资金流水')
        return redirect('finance_pages:fund_flow_management')
    
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, '请选择要导入的文件')
            return redirect('finance_pages:fund_flow_management')
        
        upload_file = request.FILES['file']
        if not upload_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, '请上传Excel文件（.xls或.xlsx格式）')
            return redirect('finance_pages:fund_flow_management')
        
        try:
            from django.db import transaction
            from datetime import datetime
            workbook = load_workbook(upload_file, data_only=True)
            worksheet = workbook.active
            
            # 读取表头
            headers = [cell.value for cell in worksheet[1]]
            header_map = {str(h).strip(): i for i, h in enumerate(headers) if h}
            
            # 检查必填列
            required_columns = ['发生日期', '流水类型', '金额', '账户名称', '摘要']
            missing_columns = [col for col in required_columns if col not in header_map]
            if missing_columns:
                messages.error(request, f'缺少必填列：{", ".join(missing_columns)}')
                return redirect('finance_pages:fund_flow_management')
            
            success_count = 0
            error_count = 0
            errors = []
            current_year = timezone.now().year
            
            # 读取数据行
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                # 跳过空行
                if not row or not any(row):
                    continue
                
                try:
                    # 解析日期
                    flow_date_str = str(row[header_map['发生日期']]).strip() if row[header_map['发生日期']] else None
                    if not flow_date_str:
                        error_count += 1
                        errors.append(f'第{row_idx}行：发生日期不能为空')
                        continue
                    
                    try:
                        if isinstance(row[header_map['发生日期']], datetime):
                            flow_date = row[header_map['发生日期']].date()
                        else:
                            flow_date = datetime.strptime(flow_date_str, '%Y-%m-%d').date()
                    except (ValueError, TypeError):
                        error_count += 1
                        errors.append(f'第{row_idx}行：发生日期格式错误，应为YYYY-MM-DD')
                        continue
                    
                    # 解析流水类型
                    flow_type_str = str(row[header_map['流水类型']]).strip() if row[header_map['流水类型']] else None
                    type_map = {
                        'income': 'income', '收入': 'income',
                        'expense': 'expense', '支出': 'expense',
                        'transfer': 'transfer', '转账': 'transfer',
                    }
                    flow_type = type_map.get(flow_type_str.lower(), flow_type_str)
                    if flow_type not in dict(FundFlow.TYPE_CHOICES):
                        error_count += 1
                        errors.append(f'第{row_idx}行：流水类型无效')
                        continue
                    
                    # 解析金额
                    amount_str = str(row[header_map['金额']]).strip() if row[header_map['金额']] else None
                    if not amount_str:
                        error_count += 1
                        errors.append(f'第{row_idx}行：金额不能为空')
                        continue
                    try:
                        amount = Decimal(amount_str)
                        if amount <= 0:
                            error_count += 1
                            errors.append(f'第{row_idx}行：金额必须大于0')
                            continue
                    except (ValueError, InvalidOperation):
                        error_count += 1
                        errors.append(f'第{row_idx}行：金额格式错误')
                        continue
                    
                    # 解析其他字段
                    account_name = str(row[header_map['账户名称']]).strip() if row[header_map['账户名称']] else None
                    if not account_name:
                        error_count += 1
                        errors.append(f'第{row_idx}行：账户名称不能为空')
                        continue
                    
                    counterparty = str(row[header_map.get('对方单位', -1)]).strip() if header_map.get('对方单位', -1) >= 0 and row[header_map.get('对方单位', -1)] else ''
                    summary = str(row[header_map['摘要']]).strip() if row[header_map['摘要']] else None
                    if not summary:
                        error_count += 1
                        errors.append(f'第{row_idx}行：摘要不能为空')
                        continue
                    
                    project_number = str(row[header_map.get('关联项目编号', -1)]).strip() if header_map.get('关联项目编号', -1) >= 0 and row[header_map.get('关联项目编号', -1)] else None
                    
                    # 查找关联项目
                    project = None
                    if project_number:
                        try:
                            from backend.apps.production_management.models import Project
                            project = Project.objects.get(project_number=project_number)
                        except Project.DoesNotExist:
                            error_count += 1
                            errors.append(f'第{row_idx}行：项目编号 {project_number} 不存在')
                            continue
                    
                    # 生成流水号
                    max_flow = FundFlow.objects.filter(
                        flow_number__startswith=f'FLOW-{current_year}-'
                    ).order_by('-flow_number').first()
                    
                    if max_flow:
                        try:
                            seq = int(max_flow.flow_number.split('-')[-1]) + 1
                        except (ValueError, IndexError):
                            seq = 1
                    else:
                        seq = 1
                    
                    flow_number = f'FLOW-{current_year}-{seq:04d}'
                    
                    # 创建资金流水
                    with transaction.atomic():
                        FundFlow.objects.create(
                            flow_number=flow_number,
                            flow_date=flow_date,
                            flow_type=flow_type,
                            amount=amount,
                            account_name=account_name,
                            counterparty=counterparty,
                            summary=summary,
                            project=project,
                            created_by=request.user,
                        )
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'第{row_idx}行：{str(e)}')
            
            if success_count > 0:
                messages.success(request, f'成功导入 {success_count} 条资金流水')
            if error_count > 0:
                error_msg = f'导入失败 {error_count} 条记录'
                if len(errors) <= 10:
                    error_msg += '：' + '；'.join(errors)
                else:
                    error_msg += f'：前10个错误：' + '；'.join(errors[:10])
                messages.warning(request, error_msg)
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('导入资金流水失败: %s', str(e))
            messages.error(request, f'导入失败：{str(e)}')
    
    return redirect('finance_pages:fund_flow_management')


@login_required
def fund_flow_export(request):
    """导出资金流水列表为Excel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.fund_flow.view', permission_codes):
        messages.error(request, '您没有权限导出资金流水')
        return redirect('finance_pages:fund_flow_management')
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    flow_type = request.GET.get('flow_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    fund_flows = FundFlow.objects.select_related('project', 'voucher', 'created_by').order_by('-flow_date', '-flow_number')
    
    if search:
        fund_flows = fund_flows.filter(
            Q(flow_number__icontains=search) |
            Q(account_name__icontains=search) |
            Q(counterparty__icontains=search) |
            Q(summary__icontains=search)
        )
    if flow_type:
        fund_flows = fund_flows.filter(flow_type=flow_type)
    if date_from:
        fund_flows = fund_flows.filter(flow_date__gte=date_from)
    if date_to:
        fund_flows = fund_flows.filter(flow_date__lte=date_to)
    
    # 创建Excel工作簿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '资金流水'
    
    headers = ['流水号', '发生日期', '流水类型', '金额', '账户名称', '对方单位', '摘要', '关联项目', '关联凭证', '创建人', '创建时间']
    worksheet.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加数据
    type_dict = dict(FundFlow.TYPE_CHOICES)
    for flow in fund_flows:
        row = [
            flow.flow_number,
            flow.flow_date.strftime('%Y-%m-%d') if flow.flow_date else '',
            type_dict.get(flow.flow_type, flow.flow_type),
            float(flow.amount),
            flow.account_name,
            flow.counterparty or '',
            flow.summary,
            flow.project.project_number if flow.project else '',
            flow.voucher.voucher_number if flow.voucher else '',
            flow.created_by.get_full_name() if flow.created_by else '',
            flow.created_time.strftime('%Y-%m-%d %H:%M') if flow.created_time else '',
        ]
        worksheet.append(row)
    
    # 调整列宽
    column_widths = [18, 12, 10, 12, 15, 20, 30, 15, 18, 12, 18]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('资金流水_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def voucher_detail(request, voucher_id):
    """凭证详情"""
    voucher = get_object_or_404(Voucher.objects.select_related('preparer', 'reviewer', 'posted_by'), id=voucher_id)
    
    # 获取凭证分录
    try:
        entries = voucher.entries.select_related('account_subject').order_by('line_number')
    except Exception:
        entries = []
    
    # 检查权限
    permission_codes = get_user_permission_codes(request.user)
    can_review = _permission_granted('financial_management.voucher.review', permission_codes)
    can_post = _permission_granted('financial_management.voucher.post', permission_codes)
    can_edit = _permission_granted('financial_management.voucher.manage', permission_codes) and voucher.status != 'posted'
    can_delete = _permission_granted('financial_management.voucher.manage', permission_codes) and voucher.status == 'draft'
    can_print = True  # 所有用户都可以打印
    
    # 计算统计信息
    entry_count = len(entries)
    debit_count = sum(1 for e in entries if e.debit_amount and e.debit_amount > 0)
    credit_count = sum(1 for e in entries if e.credit_amount and e.credit_amount > 0)
    
    # 检查借贷平衡
    is_balanced = voucher.total_debit == voucher.total_credit
    balance_diff = abs(voucher.total_debit - voucher.total_credit)
    
    # 获取关联的总账记录数
    try:
        ledger_count = Ledger.objects.filter(
            period_year=voucher.voucher_date.year if voucher.voucher_date else None,
            period_month=voucher.voucher_date.month if voucher.voucher_date else None,
        ).filter(
            account_subject__in=[e.account_subject for e in entries if e.account_subject]
        ).count() if voucher.status == 'posted' else 0
    except Exception:
        ledger_count = 0
    
    # 构建统计卡片
    summary_cards = []
    
    if voucher.status == 'posted':
        summary_cards.append({
            "label": "总账记录",
            "value": ledger_count,
            "hint": "已生成的总账记录数"
        })
    
    context = _context(
        f"凭证详情 - {voucher.voucher_number}",
        "📝",
        f"查看记账凭证 {voucher.voucher_number} 的详细信息和分录",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
        'entries': entries,
        'can_review': can_review,
        'can_post': can_post,
        'can_edit': can_edit,
        'can_delete': can_delete,
        'can_print': can_print,
        'is_balanced': is_balanced,
        'balance_diff': balance_diff,
        'entry_count': entry_count,
    })
    return render(request, "financial_management/voucher_detail.html", context)


@login_required
def voucher_entry_add(request, voucher_id):
    """快速添加凭证分录（AJAX）"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        return JsonResponse({'success': False, 'error': '您没有权限添加分录'}, status=403)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    
    # 已过账的凭证不能添加分录
    if voucher.status == 'posted':
        return JsonResponse({'success': False, 'error': '已过账的凭证不能添加分录'}, status=400)
    
    if request.method == 'POST':
        try:
            account_subject_id = request.POST.get('account_subject_id')
            summary = request.POST.get('summary', '')
            debit_amount_str = request.POST.get('debit_amount', '0')
            credit_amount_str = request.POST.get('credit_amount', '0')
            
            if not account_subject_id:
                return JsonResponse({'success': False, 'error': '请选择会计科目'}, status=400)
            
            account_subject = AccountSubject.objects.get(id=account_subject_id)
            
            # 计算下一个行号
            max_line = voucher.entries.aggregate(max_line=Max('line_number'))['max_line'] or 0
            line_number = max_line + 1
            
            # 创建分录
            entry = VoucherEntry.objects.create(
                voucher=voucher,
                line_number=line_number,
                account_subject=account_subject,
                summary=summary,
                debit_amount=Decimal(debit_amount_str) if debit_amount_str else Decimal('0.00'),
                credit_amount=Decimal(credit_amount_str) if credit_amount_str else Decimal('0.00'),
            )
            
            # 重新计算凭证合计
            total_debit = voucher.entries.aggregate(total=Sum('debit_amount'))['total'] or Decimal('0.00')
            total_credit = voucher.entries.aggregate(total=Sum('credit_amount'))['total'] or Decimal('0.00')
            voucher.total_debit = total_debit
            voucher.total_credit = total_credit
            voucher.save()
            
            return JsonResponse({
                'success': True,
                'entry': {
                    'id': entry.id,
                    'line_number': entry.line_number,
                    'account_subject': entry.account_subject.code + ' - ' + entry.account_subject.name,
                    'summary': entry.summary,
                    'debit_amount': str(entry.debit_amount),
                    'credit_amount': str(entry.credit_amount),
                },
                'voucher': {
                    'total_debit': str(voucher.total_debit),
                    'total_credit': str(voucher.total_credit),
                }
            })
        except AccountSubject.DoesNotExist:
            return JsonResponse({'success': False, 'error': '会计科目不存在'}, status=400)
        except (ValueError, InvalidOperation) as e:
            return JsonResponse({'success': False, 'error': f'金额格式错误：{str(e)}'}, status=400)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('添加分录失败: %s', str(e))
            return JsonResponse({'success': False, 'error': f'添加分录失败：{str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)


@login_required
def voucher_entry_update(request, voucher_id, entry_id):
    """快速编辑凭证分录（AJAX）"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        return JsonResponse({'success': False, 'error': '您没有权限编辑分录'}, status=403)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    entry = get_object_or_404(VoucherEntry, id=entry_id, voucher=voucher)
    
    # 已过账的凭证不能编辑分录
    if voucher.status == 'posted':
        return JsonResponse({'success': False, 'error': '已过账的凭证不能编辑分录'}, status=400)
    
    if request.method == 'POST':
        try:
            account_subject_id = request.POST.get('account_subject_id')
            summary = request.POST.get('summary', '')
            debit_amount_str = request.POST.get('debit_amount', '0')
            credit_amount_str = request.POST.get('credit_amount', '0')
            line_number_str = request.POST.get('line_number', '')
            
            if account_subject_id:
                account_subject = AccountSubject.objects.get(id=account_subject_id)
                entry.account_subject = account_subject
            
            if summary:
                entry.summary = summary
            
            if debit_amount_str:
                entry.debit_amount = Decimal(debit_amount_str) if debit_amount_str else Decimal('0.00')
            
            if credit_amount_str:
                entry.credit_amount = Decimal(credit_amount_str) if credit_amount_str else Decimal('0.00')
            
            if line_number_str:
                line_number = int(line_number_str)
                # 检查行号是否冲突
                existing_entry = voucher.entries.filter(line_number=line_number).exclude(id=entry_id).first()
                if existing_entry:
                    return JsonResponse({'success': False, 'error': f'行号 {line_number} 已被使用'}, status=400)
                entry.line_number = line_number
            
            entry.save()
            
            # 重新计算凭证合计
            total_debit = voucher.entries.aggregate(total=Sum('debit_amount'))['total'] or Decimal('0.00')
            total_credit = voucher.entries.aggregate(total=Sum('credit_amount'))['total'] or Decimal('0.00')
            voucher.total_debit = total_debit
            voucher.total_credit = total_credit
            voucher.save()
            
            return JsonResponse({
                'success': True,
                'entry': {
                    'id': entry.id,
                    'line_number': entry.line_number,
                    'account_subject': entry.account_subject.code + ' - ' + entry.account_subject.name,
                    'summary': entry.summary,
                    'debit_amount': str(entry.debit_amount),
                    'credit_amount': str(entry.credit_amount),
                },
                'voucher': {
                    'total_debit': str(voucher.total_debit),
                    'total_credit': str(voucher.total_credit),
                }
            })
        except AccountSubject.DoesNotExist:
            return JsonResponse({'success': False, 'error': '会计科目不存在'}, status=400)
        except (ValueError, InvalidOperation) as e:
            return JsonResponse({'success': False, 'error': f'金额格式错误：{str(e)}'}, status=400)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('编辑分录失败: %s', str(e))
            return JsonResponse({'success': False, 'error': f'编辑分录失败：{str(e)}'}, status=500)
    
    # GET请求，返回分录信息
    return JsonResponse({
        'success': True,
        'entry': {
            'id': entry.id,
            'line_number': entry.line_number,
            'account_subject_id': entry.account_subject.id,
            'account_subject': entry.account_subject.code + ' - ' + entry.account_subject.name,
            'summary': entry.summary,
            'debit_amount': str(entry.debit_amount),
            'credit_amount': str(entry.credit_amount),
        }
    })


@login_required
def voucher_entry_delete(request, voucher_id, entry_id):
    """快速删除凭证分录（AJAX）"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        return JsonResponse({'success': False, 'error': '您没有权限删除分录'}, status=403)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    entry = get_object_or_404(VoucherEntry, id=entry_id, voucher=voucher)
    
    # 已过账的凭证不能删除分录
    if voucher.status == 'posted':
        return JsonResponse({'success': False, 'error': '已过账的凭证不能删除分录'}, status=400)
    
    if request.method == 'POST':
        try:
            entry.delete()
            
            # 重新计算凭证合计
            total_debit = voucher.entries.aggregate(total=Sum('debit_amount'))['total'] or Decimal('0.00')
            total_credit = voucher.entries.aggregate(total=Sum('credit_amount'))['total'] or Decimal('0.00')
            voucher.total_debit = total_debit
            voucher.total_credit = total_credit
            voucher.save()
            
            return JsonResponse({
                'success': True,
                'voucher': {
                    'total_debit': str(voucher.total_debit),
                    'total_credit': str(voucher.total_credit),
                }
            })
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除分录失败: %s', str(e))
            return JsonResponse({'success': False, 'error': f'删除分录失败：{str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)


@login_required
def voucher_validate(request, voucher_id):
    """校验凭证数据（AJAX）"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.view', permission_codes):
        return HttpResponse('您没有权限校验凭证', status=403)
    
    voucher = get_object_or_404(Voucher.objects.prefetch_related('entries'), id=voucher_id)
    
    errors = []
    warnings = []
    
    # 检查是否有分录
    if not voucher.entries.exists():
        errors.append('凭证至少需要一条分录')
    
    # 检查借贷平衡
    if voucher.total_debit != voucher.total_credit:
        diff = abs(voucher.total_debit - voucher.total_credit)
        errors.append(f'借贷不平衡，差额：{diff:,.2f}')
    
    # 检查分录中的科目是否有效
    for entry in voucher.entries.all():
        if not entry.account_subject.is_active:
            warnings.append(f'第{entry.line_number}行：科目 {entry.account_subject.code} 已停用')
        
        # 检查借方和贷方不能同时有值
        if entry.debit_amount > 0 and entry.credit_amount > 0:
            errors.append(f'第{entry.line_number}行：借方和贷方不能同时有金额')
        
        # 检查借方和贷方不能同时为0
        if entry.debit_amount == 0 and entry.credit_amount == 0:
            warnings.append(f'第{entry.line_number}行：借方和贷方金额都为0')
    
    # 检查行号是否连续
    line_numbers = sorted([e.line_number for e in voucher.entries.all()])
    if line_numbers:
        expected_lines = list(range(1, len(line_numbers) + 1))
        if line_numbers != expected_lines:
            warnings.append('分录行号不连续')
    
    result = {
        'valid': len(errors) == 0,
        'errors': errors,
        'warnings': warnings,
    }
    
    return JsonResponse(result)


@login_required
def voucher_copy(request, voucher_id):
    """复制凭证（创建新凭证）"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.create', permission_codes):
        messages.error(request, '您没有权限创建凭证')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    source_voucher = get_object_or_404(Voucher.objects.prefetch_related('entries'), id=voucher_id)
    
    if request.method == 'POST':
        try:
            from django.db import transaction
            with transaction.atomic():
                # 创建新凭证
                new_voucher = Voucher.objects.create(
                    voucher_date=request.POST.get('voucher_date', timezone.now().date()),
                    preparer=request.user,
                    status='draft',
                    attachment_count=0,
                    total_debit=source_voucher.total_debit,
                    total_credit=source_voucher.total_credit,
                    notes=f"复制自 {source_voucher.voucher_number}" + (f"\n{source_voucher.notes}" if source_voucher.notes else ""),
                )
                
                # 自动生成凭证字号
                new_voucher.voucher_number = _generate_voucher_number(new_voucher.voucher_date)
                new_voucher.save()
                
                # 复制凭证分录
                for entry in source_voucher.entries.all():
                    VoucherEntry.objects.create(
                        voucher=new_voucher,
                        line_number=entry.line_number,
                        account_subject=entry.account_subject,
                        summary=entry.summary,
                        debit_amount=entry.debit_amount,
                        credit_amount=entry.credit_amount,
                    )
                
                messages.success(request, f'凭证已复制，新凭证号：{new_voucher.voucher_number}')
                return redirect('finance_pages:voucher_detail', voucher_id=new_voucher.id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('复制凭证失败: %s', str(e))
            messages.error(request, f'复制凭证失败：{str(e)}')
            return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    # GET请求，显示确认页面
    context = _context(
        f"复制凭证 - {source_voucher.voucher_number}",
        "📋",
        f"复制凭证 {source_voucher.voucher_number} 创建新凭证",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'source_voucher': source_voucher,
        'entries': source_voucher.entries.select_related('account_subject').order_by('line_number'),
    })
    return render(request, "financial_management/voucher_copy.html", context)


@login_required
def voucher_print(request, voucher_id):
    """打印凭证（PDF格式）"""
    if not REPORTLAB_AVAILABLE:
        messages.error(request, 'PDF打印功能需要安装reportlab库')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher.objects.select_related('preparer', 'reviewer', 'posted_by'), id=voucher_id)
    
    # 获取凭证分录
    try:
        entries = voucher.entries.select_related('account_subject').order_by('line_number')
    except Exception:
        entries = []
    
    # 创建PDF响应
    response = HttpResponse(content_type='application/pdf')
    filename = f'凭证_{voucher.voucher_number}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # 创建PDF文档
    doc = SimpleDocTemplate(response, pagesize=A4, 
                           rightMargin=20*mm, leftMargin=20*mm,
                           topMargin=20*mm, bottomMargin=20*mm)
    
    # 构建内容
    story = []
    styles = getSampleStyleSheet()
    
    # 标题样式
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=12,
        alignment=1,  # 居中
    )
    
    # 添加标题
    story.append(Paragraph('记账凭证', title_style))
    story.append(Spacer(1, 6*mm))
    
    # 凭证基本信息表格
    voucher_info_data = [
        ['凭证字号', voucher.voucher_number, '凭证日期', voucher.voucher_date.strftime('%Y-%m-%d') if voucher.voucher_date else ''],
        ['制单人', voucher.preparer.get_full_name() if voucher.preparer else '', '附件数', str(voucher.attachment_count)],
    ]
    
    if voucher.reviewer:
        voucher_info_data.append(['审核人', voucher.reviewer.get_full_name(), '审核时间', voucher.reviewed_time.strftime('%Y-%m-%d %H:%M') if voucher.reviewed_time else ''])
    
    if voucher.posted_by:
        voucher_info_data.append(['过账人', voucher.posted_by.get_full_name(), '过账时间', voucher.posted_time.strftime('%Y-%m-%d %H:%M') if voucher.posted_time else ''])
    
    voucher_info_table = Table(voucher_info_data, colWidths=[40*mm, 50*mm, 40*mm, 50*mm])
    voucher_info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(voucher_info_table)
    story.append(Spacer(1, 6*mm))
    
    # 凭证分录表格
    entry_headers = ['行号', '会计科目', '摘要', '借方金额', '贷方金额']
    entry_data = [entry_headers]
    
    for entry in entries:
        entry_data.append([
            str(entry.line_number),
            f"{entry.account_subject.code} {entry.account_subject.name}" if entry.account_subject else '',
            entry.summary,
            f"{entry.debit_amount:,.2f}" if entry.debit_amount else '0.00',
            f"{entry.credit_amount:,.2f}" if entry.credit_amount else '0.00',
        ])
    
    # 添加合计行
    entry_data.append([
        '',
        '',
        '合计',
        f"{voucher.total_debit:,.2f}",
        f"{voucher.total_credit:,.2f}",
    ])
    
    entry_table = Table(entry_data, colWidths=[15*mm, 50*mm, 60*mm, 30*mm, 30*mm])
    entry_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'SimSun'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('FONTNAME', (0, 0), (-1, 0), 'SimSun'),
        ('BOLD', (0, 0), (-1, 0), True),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
        ('BOLD', (0, -1), (-1, -1), True),
    ]))
    story.append(entry_table)
    story.append(Spacer(1, 6*mm))
    
    # 备注信息
    if voucher.notes:
        story.append(Paragraph(f'<b>备注：</b>{voucher.notes}', styles['Normal']))
        story.append(Spacer(1, 6*mm))
    
    # 状态信息
    status_dict = dict(Voucher.STATUS_CHOICES)
    status_text = status_dict.get(voucher.status, voucher.status)
    story.append(Paragraph(f'<b>状态：</b>{status_text}', styles['Normal']))
    
    # 生成PDF
    doc.build(story)
    return response


@login_required
def voucher_submit(request, voucher_id):
    """提交凭证（草稿->已提交）"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        messages.error(request, '您没有权限提交凭证')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    
    # 检查状态
    if voucher.status != 'draft':
        messages.error(request, f'只有草稿状态的凭证才能提交，当前状态：{voucher.get_status_display()}')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # 检查借贷是否平衡
    if voucher.total_debit != voucher.total_credit:
        messages.error(request, f'凭证借贷不平衡（借方：{voucher.total_debit}，贷方：{voucher.total_credit}），不能提交')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # 检查是否有分录
    if not voucher.entries.exists():
        messages.error(request, '凭证至少需要一条分录才能提交')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        voucher.status = 'submitted'
        voucher.save()
        messages.success(request, f'凭证 {voucher.voucher_number} 已提交，等待审核')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # GET请求，显示确认页面
    context = _context(
        f"提交凭证 - {voucher.voucher_number}",
        "📤",
        f"提交记账凭证 {voucher.voucher_number} 等待审核",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
        'entries': voucher.entries.select_related('account_subject').order_by('line_number'),
    })
    return render(request, "financial_management/voucher_submit.html", context)


@login_required
def voucher_withdraw(request, voucher_id):
    """撤回凭证（已提交->草稿）"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        messages.error(request, '您没有权限撤回凭证')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    
    # 检查状态
    if voucher.status != 'submitted':
        messages.error(request, f'只有已提交状态的凭证才能撤回，当前状态：{voucher.get_status_display()}')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # 检查权限：只有制单人才能撤回
    if voucher.preparer != request.user:
        messages.error(request, '只有制单人才能撤回凭证')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        voucher.status = 'draft'
        voucher.save()
        messages.success(request, f'凭证 {voucher.voucher_number} 已撤回，可以继续编辑')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # GET请求，显示确认页面
    context = _context(
        f"撤回凭证 - {voucher.voucher_number}",
        "↩️",
        f"撤回记账凭证 {voucher.voucher_number} 返回草稿状态",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
    })
    return render(request, "financial_management/voucher_withdraw.html", context)


@login_required
def voucher_approve(request, voucher_id):
    """审核凭证"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.review', permission_codes):
        messages.error(request, '您没有权限审核凭证')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher, id=voucher_id)
    
    # 检查状态
    if voucher.status not in ['submitted', 'draft']:
        messages.error(request, f'凭证状态为 {voucher.get_status_display()}，不能审核')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # 检查借贷是否平衡
    if voucher.total_debit != voucher.total_credit:
        messages.error(request, f'凭证借贷不平衡（借方：{voucher.total_debit}，贷方：{voucher.total_credit}），不能审核')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        action = request.POST.get('action', 'approve')
        
        if action == 'approve':
            # 审核通过
            voucher.status = 'approved'
            voucher.reviewer = request.user
            voucher.reviewed_time = timezone.now()
            voucher.save()
            messages.success(request, f'凭证 {voucher.voucher_number} 审核通过')
        elif action == 'reject':
            # 审核拒绝
            voucher.status = 'rejected'
            voucher.reviewer = request.user
            voucher.reviewed_time = timezone.now()
            voucher.save()
            messages.success(request, f'凭证 {voucher.voucher_number} 已拒绝')
        
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # GET请求，显示确认页面
    context = _context(
        f"审核凭证 - {voucher.voucher_number}",
        "✅",
        f"审核记账凭证 {voucher.voucher_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
    })
    return render(request, "financial_management/voucher_approve.html", context)


@login_required
def voucher_post(request, voucher_id):
    """过账凭证"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.post', permission_codes):
        messages.error(request, '您没有权限过账凭证')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher.objects.prefetch_related('entries'), id=voucher_id)
    
    # 检查状态
    if voucher.status != 'approved':
        messages.error(request, f'只有已审核的凭证才能过账，当前状态：{voucher.get_status_display()}')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # 检查借贷是否平衡
    if voucher.total_debit != voucher.total_credit:
        messages.error(request, f'凭证借贷不平衡（借方：{voucher.total_debit}，贷方：{voucher.total_credit}），不能过账')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        from django.db import transaction
        
        try:
            with transaction.atomic():
                # 更新凭证状态
                voucher.status = 'posted'
                voucher.posted_by = request.user
                voucher.posted_time = timezone.now()
                voucher.save()
                
                # 生成总账记录
                period_year = voucher.voucher_date.year
                period_month = voucher.voucher_date.month
                
                for entry in voucher.entries.all():
                    # 获取或创建总账记录
                    ledger, created = Ledger.objects.get_or_create(
                        account_subject=entry.account_subject,
                        period_year=period_year,
                        period_month=period_month,
                        period_date=voucher.voucher_date,
                        defaults={
                            'opening_balance': Decimal('0.00'),
                            'period_debit': Decimal('0.00'),
                            'period_credit': Decimal('0.00'),
                            'closing_balance': Decimal('0.00'),
                        }
                    )
                    
                    # 更新总账金额
                    ledger.period_debit += entry.debit_amount or Decimal('0.00')
                    ledger.period_credit += entry.credit_amount or Decimal('0.00')
                    
                    # 计算期末余额（根据科目余额方向）
                    if entry.account_subject.direction == 'debit':
                        # 借方科目：期初余额 + 借方 - 贷方
                        ledger.closing_balance = ledger.opening_balance + ledger.period_debit - ledger.period_credit
                    else:
                        # 贷方科目：期初余额 + 贷方 - 借方
                        ledger.closing_balance = ledger.opening_balance + ledger.period_credit - ledger.period_debit
                    
                    ledger.save()
                
                messages.success(request, f'凭证 {voucher.voucher_number} 过账成功，已生成总账记录')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('过账失败: %s', str(e))
            messages.error(request, f'过账失败：{str(e)}')
        
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # GET请求，显示确认页面
    context = _context(
        f"过账凭证 - {voucher.voucher_number}",
        "📖",
        f"将凭证 {voucher.voucher_number} 过账到总账",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
        'entries': voucher.entries.select_related('account_subject').order_by('line_number'),
    })
    return render(request, "financial_management/voucher_post.html", context)


@login_required
def voucher_unpost(request, voucher_id):
    """反过账凭证（已过账->已审核）"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.voucher.post', permission_codes):
        messages.error(request, '您没有权限反过账凭证')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    voucher = get_object_or_404(Voucher.objects.prefetch_related('entries'), id=voucher_id)
    
    # 检查状态
    if voucher.status != 'posted':
        messages.error(request, f'只有已过账的凭证才能反过账，当前状态：{voucher.get_status_display()}')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    if request.method == 'POST':
        from django.db import transaction
        
        try:
            with transaction.atomic():
                # 删除总账记录（回滚过账操作）
                period_year = voucher.voucher_date.year
                period_month = voucher.voucher_date.month
                
                for entry in voucher.entries.all():
                    try:
                        ledger = Ledger.objects.get(
                            account_subject=entry.account_subject,
                            period_year=period_year,
                            period_month=period_month,
                            period_date=voucher.voucher_date,
                        )
                        
                        # 回滚总账金额
                        ledger.period_debit -= entry.debit_amount or Decimal('0.00')
                        ledger.period_credit -= entry.credit_amount or Decimal('0.00')
                        
                        # 重新计算期末余额
                        if entry.account_subject.direction == 'debit':
                            ledger.closing_balance = ledger.opening_balance + ledger.period_debit - ledger.period_credit
                        else:
                            ledger.closing_balance = ledger.opening_balance + ledger.period_credit - ledger.period_debit
                        
                        # 如果本期借贷都为0，可以删除该总账记录
                        if ledger.period_debit == Decimal('0.00') and ledger.period_credit == Decimal('0.00'):
                            ledger.delete()
                        else:
                            ledger.save()
                    except Ledger.DoesNotExist:
                        # 总账记录不存在，跳过
                        pass
                
                # 更新凭证状态
                voucher.status = 'approved'
                voucher.posted_by = None
                voucher.posted_time = None
                voucher.save()
                
                messages.success(request, f'凭证 {voucher.voucher_number} 反过账成功，已回滚总账记录')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('反过账失败: %s', str(e))
            messages.error(request, f'反过账失败：{str(e)}')
        
        return redirect('finance_pages:voucher_detail', voucher_id=voucher.id)
    
    # GET请求，显示确认页面
    context = _context(
        f"反过账凭证 - {voucher.voucher_number}",
        "↩️",
        f"将凭证 {voucher.voucher_number} 反过账，回滚总账记录",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
        'entries': voucher.entries.select_related('account_subject').order_by('line_number'),
    })
    return render(request, "financial_management/voucher_unpost.html", context)


@login_required
def voucher_delete(request, voucher_id):
    """删除凭证"""
    voucher = get_object_or_404(Voucher, id=voucher_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.voucher.manage', permission_codes):
        messages.error(request, '您没有权限删除凭证')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    # 只有草稿状态的凭证可以删除
    if voucher.status != 'draft':
        messages.error(request, f'只能删除草稿状态的凭证，当前状态：{voucher.get_status_display()}')
        return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    if request.method == 'POST':
        try:
            voucher_number = voucher.voucher_number
            voucher.delete()
            messages.success(request, f'凭证 {voucher_number} 已删除')
            return redirect('finance_pages:voucher_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除凭证失败: %s', str(e))
            messages.error(request, f'删除凭证失败：{str(e)}')
            return redirect('finance_pages:voucher_detail', voucher_id=voucher_id)
    
    context = _context(
        f"删除凭证 - {voucher.voucher_number}",
        "🗑️",
        f"确认删除凭证：{voucher.voucher_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'voucher': voucher,
    })
    return render(request, "financial_management/voucher_delete.html", context)


@login_required
def budget_detail(request, budget_id):
    """预算详情"""
    budget = get_object_or_404(Budget.objects.select_related('department', 'account_subject', 'approver', 'created_by'), id=budget_id)
    
    # 计算使用率
    usage_rate = 0
    if budget.budget_amount > 0:
        usage_rate = (budget.used_amount / budget.budget_amount) * 100
    
    # 检查权限
    permission_codes = get_user_permission_codes(request.user)
    can_approve = _permission_granted('financial_management.budget.approve', permission_codes)
    
    context = _context(
        f"预算详情 - {budget.budget_number}",
        "💰",
        f"查看预算 {budget.name} 的详细信息",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'budget': budget,
        'usage_rate': usage_rate,
        'can_approve': can_approve,
    })
    return render(request, "financial_management/budget_detail.html", context)


@login_required
def budget_delete(request, budget_id):
    """删除预算"""
    budget = get_object_or_404(Budget, id=budget_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.budget.manage', permission_codes):
        messages.error(request, '您没有权限删除预算')
        return redirect('finance_pages:budget_detail', budget_id=budget_id)
    
    # 只有草稿状态的预算可以删除
    if budget.status != 'draft':
        messages.error(request, f'只能删除草稿状态的预算，当前状态：{budget.get_status_display()}')
        return redirect('finance_pages:budget_detail', budget_id=budget_id)
    
    if request.method == 'POST':
        try:
            budget_number = budget.budget_number
            budget.delete()
            messages.success(request, f'预算 {budget_number} 已删除')
            return redirect('finance_pages:budget_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除预算失败: %s', str(e))
            messages.error(request, f'删除预算失败：{str(e)}')
            return redirect('finance_pages:budget_detail', budget_id=budget_id)
    
    context = _context(
        f"删除预算 - {budget.budget_number}",
        "🗑️",
        f"确认删除预算：{budget.name}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'budget': budget,
    })
    return render(request, "financial_management/budget_delete.html", context)


@login_required
def budget_approve(request, budget_id):
    """审批预算"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.budget.approve', permission_codes):
        messages.error(request, '您没有权限审批预算')
        return redirect('finance_pages:budget_detail', budget_id=budget_id)
    
    budget = get_object_or_404(Budget.objects.select_related('department', 'account_subject', 'created_by'), id=budget_id)
    
    # 检查状态
    if budget.status != 'draft':
        messages.error(request, f'预算状态为 {budget.get_status_display()}，不能审批')
        return redirect('finance_pages:budget_detail', budget_id=budget.id)
    
    if request.method == 'POST':
        action = request.POST.get('action', 'approve')
        
        if action == 'approve':
            # 审批通过
            budget.status = 'approved'
            budget.approver = request.user
            budget.approved_time = timezone.now()
            # 自动计算剩余金额
            budget.remaining_amount = budget.budget_amount - budget.used_amount
            budget.save()
            messages.success(request, f'预算 {budget.budget_number} 审批通过')
        elif action == 'reject':
            # 审批拒绝（取消）
            budget.status = 'cancelled'
            budget.approver = request.user
            budget.approved_time = timezone.now()
            budget.save()
            messages.success(request, f'预算 {budget.budget_number} 已取消')
        
        return redirect('finance_pages:budget_detail', budget_id=budget.id)
    
    # GET请求，显示确认页面
    context = _context(
        f"审批预算 - {budget.budget_number}",
        "✅",
        f"审批预算 {budget.name}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'budget': budget,
    })
    return render(request, "financial_management/budget_approve.html", context)


@login_required
def invoice_detail(request, invoice_id):
    """发票详情"""
    invoice = get_object_or_404(Invoice.objects.select_related('verified_by', 'created_by'), id=invoice_id)
    
    # 检查权限
    permission_codes = get_user_permission_codes(request.user)
    can_verify = _permission_granted('financial_management.invoice.manage', permission_codes) and invoice.status == 'issued'
    can_edit = _permission_granted('financial_management.invoice.manage', permission_codes) and invoice.status != 'verified'
    
    context = _context(
        f"发票详情 - {invoice.invoice_number}",
        "🧾",
        f"查看发票 {invoice.invoice_number} 的详细信息",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'invoice': invoice,
        'can_verify': can_verify,
        'can_edit': can_edit,
    })
    return render(request, "financial_management/invoice_detail.html", context)


@login_required
def invoice_verify(request, invoice_id):
    """认证发票"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.invoice.manage', permission_codes):
        messages.error(request, '您没有权限认证发票')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice_id)
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    # 检查状态
    if invoice.status != 'issued':
        messages.error(request, f'只有已开具状态的发票才能认证，当前状态：{invoice.get_status_display()}')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    
    if request.method == 'POST':
        invoice.status = 'verified'
        invoice.verified_by = request.user
        invoice.verified_time = timezone.now()
        invoice.save()
        messages.success(request, f'发票 {invoice.invoice_number} 认证成功')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    
    # GET请求，显示确认页面
    context = _context(
        f"认证发票 - {invoice.invoice_number}",
        "✅",
        f"认证发票 {invoice.invoice_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'invoice': invoice,
    })
    return render(request, "financial_management/invoice_verify.html", context)


@login_required
def invoice_cancel(request, invoice_id):
    """作废发票"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.invoice.manage', permission_codes):
        messages.error(request, '您没有权限作废发票')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice_id)
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    # 检查状态：已认证的发票不能作废
    if invoice.status == 'verified':
        messages.error(request, '已认证的发票不能作废')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    
    if request.method == 'POST':
        invoice.status = 'cancelled'
        invoice.save()
        messages.success(request, f'发票 {invoice.invoice_number} 已作废')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice.id)
    
    # GET请求，显示确认页面
    context = _context(
        f"作废发票 - {invoice.invoice_number}",
        "❌",
        f"作废发票 {invoice.invoice_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'invoice': invoice,
    })
    return render(request, "financial_management/invoice_cancel.html", context)


@login_required
def invoice_delete(request, invoice_id):
    """删除发票"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.invoice.manage', permission_codes):
        messages.error(request, '您没有权限删除发票')
        return redirect('finance_pages:invoice_detail', invoice_id=invoice_id)
    
    if request.method == 'POST':
        try:
            invoice_number = invoice.invoice_number
            invoice.delete()
            messages.success(request, f'发票 {invoice_number} 已删除')
            return redirect('finance_pages:invoice_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除发票失败: %s', str(e))
            messages.error(request, f'删除发票失败：{str(e)}')
            return redirect('finance_pages:invoice_detail', invoice_id=invoice_id)
    
    context = _context(
        f"删除发票 - {invoice.invoice_number}",
        "🗑️",
        f"确认删除发票：{invoice.invoice_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'invoice': invoice,
    })
    return render(request, "financial_management/invoice_delete.html", context)


@login_required
def account_subject_detail(request, account_subject_id):
    """会计科目详情"""
    account_subject = get_object_or_404(
        AccountSubject.objects.select_related('parent', 'created_by'),
        id=account_subject_id
    )
    
    # 获取子科目
    try:
        children = AccountSubject.objects.filter(parent=account_subject).order_by('code')
    except Exception:
        children = []
    
    # 获取使用统计
    try:
        voucher_entry_count = account_subject.voucher_entries.count()
        ledger_entry_count = account_subject.ledger_entries.count()
    except Exception:
        voucher_entry_count = 0
        ledger_entry_count = 0
    
    context = _context(
        f"会计科目详情 - {account_subject.code} {account_subject.name}",
        "📊",
        f"查看会计科目 {account_subject.code} {account_subject.name} 的详细信息",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'account_subject': account_subject,
        'children': children,
        'voucher_entry_count': voucher_entry_count,
        'ledger_entry_count': ledger_entry_count,
    })
    return render(request, "financial_management/account_subject_detail.html", context)


@login_required
def ledger_detail(request, ledger_id):
    """账簿详情"""
    ledger = get_object_or_404(
        Ledger.objects.select_related('account_subject'),
        id=ledger_id
    )
    
    # 获取同一科目的其他期间记录（最近6个月）
    try:
        related_ledgers = Ledger.objects.filter(
            account_subject=ledger.account_subject
        ).exclude(id=ledger.id).order_by('-period_date')[:6]
    except Exception:
        related_ledgers = []
    
    context = _context(
        f"账簿详情 - {ledger.account_subject.code} {ledger.period_date}",
        "📖",
        f"查看会计科目 {ledger.account_subject.code} 在 {ledger.period_date} 的账务记录",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'ledger': ledger,
        'related_ledgers': related_ledgers,
    })
    return render(request, "financial_management/ledger_detail.html", context)


@login_required
def ledger_opening_balance_setup(request):
    """设置期初余额"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.ledger.manage', permission_codes):
        messages.error(request, '您没有权限设置期初余额')
        return redirect('finance_pages:ledger_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', '1')
    
    try:
        period_year = int(period_year) if period_year else today.year
        period_month = int(period_month) if period_month else 1
    except (ValueError, TypeError):
        period_year = today.year
        period_month = 1
    
    # 获取所有启用的会计科目
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    
    if request.method == 'POST':
        from django.db import transaction
        try:
            with transaction.atomic():
                success_count = 0
                for subject in account_subjects:
                    balance_key = f'balance_{subject.id}'
                    balance_str = request.POST.get(balance_key, '0')
                    
                    try:
                        balance = Decimal(balance_str) if balance_str else Decimal('0.00')
                        
                        # 获取或创建该科目的期初余额记录（使用该期间的第一天）
                        period_date = timezone.datetime(period_year, period_month, 1).date()
                        
                        ledger, created = Ledger.objects.get_or_create(
                            account_subject=subject,
                            period_year=period_year,
                            period_month=period_month,
                            period_date=period_date,
                            defaults={
                                'opening_balance': balance,
                                'period_debit': Decimal('0.00'),
                                'period_credit': Decimal('0.00'),
                                'closing_balance': balance,
                            }
                        )
                        
                        if not created:
                            # 更新期初余额
                            ledger.opening_balance = balance
                            # 重新计算期末余额
                            if subject.direction == 'debit':
                                ledger.closing_balance = ledger.opening_balance + ledger.period_debit - ledger.period_credit
                            else:
                                ledger.closing_balance = ledger.opening_balance + ledger.period_credit - ledger.period_debit
                            ledger.save()
                        
                        success_count += 1
                    except (ValueError, InvalidOperation):
                        continue
                
                messages.success(request, f'成功设置 {success_count} 个科目的期初余额')
                return redirect(f"{reverse('finance_pages:ledger_opening_balance_setup')}?period_year={period_year}&period_month={period_month}")
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('设置期初余额失败: %s', str(e))
            messages.error(request, f'设置期初余额失败：{str(e)}')
    
    # 获取当前期间的期初余额
    period_date = timezone.datetime(period_year, period_month, 1).date()
    opening_balances = {}
    for subject in account_subjects:
        try:
            ledger = Ledger.objects.filter(
                account_subject=subject,
                period_year=period_year,
                period_month=period_month
            ).order_by('period_date').first()
            if ledger:
                opening_balances[subject.id] = ledger.opening_balance
            else:
                opening_balances[subject.id] = Decimal('0.00')
        except Exception:
            opening_balances[subject.id] = Decimal('0.00')
    
    context = _context(
        f"设置期初余额 - {period_year}年{period_month}月",
        "💰",
        f"为 {period_year}年{period_month}月 设置各科目的期初余额",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'account_subjects': account_subjects,
        'opening_balances': opening_balances,
        'period_year': period_year,
        'period_month': period_month,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/ledger_opening_balance_setup.html", context)


@login_required
def ledger_period_closing(request):
    """期末结账"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.ledger.manage', permission_codes):
        messages.error(request, '您没有权限进行期末结账')
        return redirect('finance_pages:ledger_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year) if period_year else today.year
        period_month = int(period_month) if period_month else today.month
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # 计算下一期间
    if period_month == 12:
        next_period_year = period_year + 1
        next_period_month = 1
    else:
        next_period_year = period_year
        next_period_month = period_month + 1
    
    # 检查当前期间是否有未过账的凭证
    unposted_vouchers = Voucher.objects.filter(
        voucher_date__year=period_year,
        voucher_date__month=period_month,
        status__in=['draft', 'submitted', 'approved']
    ).count()
    
    if request.method == 'POST':
        if unposted_vouchers > 0:
            messages.error(request, f'当前期间还有 {unposted_vouchers} 张凭证未过账，不能结账')
            return redirect(f"{reverse('finance_pages:ledger_period_closing')}?period_year={period_year}&period_month={period_month}")
        
        from django.db import transaction
        try:
            with transaction.atomic():
                # 获取当前期间的所有总账记录
                current_ledgers = Ledger.objects.filter(
                    period_year=period_year,
                    period_month=period_month
                ).select_related('account_subject')
                
                success_count = 0
                for ledger in current_ledgers:
                    # 获取下一期间的期初余额记录（使用下一期间的第一天）
                    next_period_date = timezone.datetime(next_period_year, next_period_month, 1).date()
                    
                    next_ledger, created = Ledger.objects.get_or_create(
                        account_subject=ledger.account_subject,
                        period_year=next_period_year,
                        period_month=next_period_month,
                        period_date=next_period_date,
                        defaults={
                            'opening_balance': ledger.closing_balance,
                            'period_debit': Decimal('0.00'),
                            'period_credit': Decimal('0.00'),
                            'closing_balance': ledger.closing_balance,
                        }
                    )
                    
                    if not created:
                        # 更新期初余额
                        next_ledger.opening_balance = ledger.closing_balance
                        # 重新计算期末余额
                        if ledger.account_subject.direction == 'debit':
                            next_ledger.closing_balance = next_ledger.opening_balance + next_ledger.period_debit - next_ledger.period_credit
                        else:
                            next_ledger.closing_balance = next_ledger.opening_balance + next_ledger.period_credit - next_ledger.period_debit
                        next_ledger.save()
                    
                    success_count += 1
                
                messages.success(request, f'成功结账：{period_year}年{period_month}月 → {next_period_year}年{next_period_month}月，共 {success_count} 个科目')
                return redirect('finance_pages:ledger_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('期末结账失败: %s', str(e))
            messages.error(request, f'期末结账失败：{str(e)}')
    
    # 获取当前期间的科目统计
    current_ledgers = Ledger.objects.filter(
        period_year=period_year,
        period_month=period_month
    ).select_related('account_subject')
    
    context = _context(
        f"期末结账 - {period_year}年{period_month}月",
        "📋",
        f"将 {period_year}年{period_month}月 的期末余额结转到 {next_period_year}年{next_period_month}月",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'period_year': period_year,
        'period_month': period_month,
        'next_period_year': next_period_year,
        'next_period_month': next_period_month,
        'unposted_vouchers': unposted_vouchers,
        'ledger_count': current_ledgers.count(),
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/ledger_period_closing.html", context)


@login_required
def subsidiary_ledger(request):
    """明细账"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.ledger.view', permission_codes):
        messages.error(request, '您没有权限查看明细账')
        return redirect('finance_pages:ledger_management')
    
    today = timezone.now().date()
    account_subject_id = request.GET.get('account_subject_id', '')
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    try:
        period_year = int(period_year) if period_year else today.year
        period_month = int(period_month) if period_month else today.month
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # 获取所有会计科目
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    
    entries = []
    account_subject = None
    opening_balance = Decimal('0.00')
    closing_balance = Decimal('0.00')
    
    if account_subject_id:
        try:
            account_subject = AccountSubject.objects.get(id=int(account_subject_id))
            
            # 获取期初余额（上期期末余额）
            if period_month > 1:
                prev_month = period_month - 1
                prev_ledger = Ledger.objects.filter(
                    account_subject=account_subject,
                    period_year=period_year,
                    period_month=prev_month
                ).order_by('-period_date').first()
                if prev_ledger:
                    opening_balance = prev_ledger.closing_balance
            else:
                prev_year = period_year - 1
                prev_ledger = Ledger.objects.filter(
                    account_subject=account_subject,
                    period_year=prev_year,
                    period_month=12
                ).order_by('-period_date').first()
                if prev_ledger:
                    opening_balance = prev_ledger.closing_balance
            
            # 获取凭证分录
            voucher_entries_query = VoucherEntry.objects.filter(
                account_subject=account_subject,
                voucher__voucher_date__year=period_year,
                voucher__voucher_date__month=period_month,
                voucher__status='posted'  # 只显示已过账的凭证
            ).select_related('voucher', 'account_subject').order_by('voucher__voucher_date', 'voucher__voucher_number', 'line_number')
            
            if date_from:
                voucher_entries_query = voucher_entries_query.filter(voucher__voucher_date__gte=date_from)
            if date_to:
                voucher_entries_query = voucher_entries_query.filter(voucher__voucher_date__lte=date_to)
            
            entries = list(voucher_entries_query)
            
            # 计算每笔分录的余额和合计
            current_balance = opening_balance
            entries_with_balance = []
            total_debit_amount = Decimal('0.00')
            total_credit_amount = Decimal('0.00')
            for entry in entries:
                if account_subject.direction == 'debit':
                    current_balance = current_balance + entry.debit_amount - entry.credit_amount
                else:
                    current_balance = current_balance + entry.credit_amount - entry.debit_amount
                entries_with_balance.append({
                    'entry': entry,
                    'balance': current_balance,
                })
                total_debit_amount += entry.debit_amount
                total_credit_amount += entry.credit_amount
            closing_balance = current_balance
            entries = entries_with_balance
            
        except (ValueError, AccountSubject.DoesNotExist):
            messages.error(request, '无效的会计科目')
    
    context = _context(
        "明细账",
        "📖",
        "查看会计科目的详细凭证分录",
        request=request,
        use_financial_nav=True
    )
    # 计算合计（如果已选择科目）
    total_debit_amount = Decimal('0.00')
    total_credit_amount = Decimal('0.00')
    if account_subject and entries:
        for item in entries:
            total_debit_amount += item['entry'].debit_amount
            total_credit_amount += item['entry'].credit_amount
    
    context.update({
        'account_subjects': account_subjects,
        'account_subject': account_subject,
        'entries': entries,
        'opening_balance': opening_balance,
        'closing_balance': closing_balance,
        'total_debit': total_debit_amount,
        'total_credit': total_credit_amount,
        'current_account_subject_id': account_subject_id,
        'current_period_year': period_year,
        'current_period_month': period_month,
        'current_date_from': date_from,
        'current_date_to': date_to,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/subsidiary_ledger.html", context)


@login_required
def account_balance_sheet(request):
    """科目余额表"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.ledger.view', permission_codes):
        messages.error(request, '您没有权限查看科目余额表')
        return redirect('finance_pages:ledger_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year) if period_year else today.year
        period_month = int(period_month) if period_month else today.month
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # 获取所有会计科目及其余额
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    balance_data = []
    
    total_opening_debit = Decimal('0.00')
    total_opening_credit = Decimal('0.00')
    total_period_debit = Decimal('0.00')
    total_period_credit = Decimal('0.00')
    total_closing_debit = Decimal('0.00')
    total_closing_credit = Decimal('0.00')
    
    for subject in account_subjects:
        # 获取该科目的总账记录
        ledger = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        ).order_by('-period_date').first()
        
        if ledger:
            opening_balance = ledger.opening_balance
            period_debit = ledger.period_debit
            period_credit = ledger.period_credit
            closing_balance = ledger.closing_balance
            
            # 根据余额方向调整显示
            if subject.direction == 'debit':
                opening_debit = opening_balance if opening_balance >= 0 else Decimal('0.00')
                opening_credit = -opening_balance if opening_balance < 0 else Decimal('0.00')
                closing_debit = closing_balance if closing_balance >= 0 else Decimal('0.00')
                closing_credit = -closing_balance if closing_balance < 0 else Decimal('0.00')
            else:
                opening_debit = -opening_balance if opening_balance < 0 else Decimal('0.00')
                opening_credit = opening_balance if opening_balance >= 0 else Decimal('0.00')
                closing_debit = -closing_balance if closing_balance < 0 else Decimal('0.00')
                closing_credit = closing_balance if closing_balance >= 0 else Decimal('0.00')
            
            balance_data.append({
                'subject': subject,
                'opening_debit': opening_debit,
                'opening_credit': opening_credit,
                'period_debit': period_debit,
                'period_credit': period_credit,
                'closing_debit': closing_debit,
                'closing_credit': closing_credit,
            })
            
            total_opening_debit += opening_debit
            total_opening_credit += opening_credit
            total_period_debit += period_debit
            total_period_credit += period_credit
            total_closing_debit += closing_debit
            total_closing_credit += closing_credit
    
    context = _context(
        "科目余额表",
        "📊",
        f"查看 {period_year}年{period_month}月 所有科目的余额情况",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'balance_data': balance_data,
        'period_year': period_year,
        'period_month': period_month,
        'total_opening_debit': total_opening_debit,
        'total_opening_credit': total_opening_credit,
        'total_period_debit': total_period_debit,
        'total_period_credit': total_period_credit,
        'total_closing_debit': total_closing_debit,
        'total_closing_credit': total_closing_credit,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/account_balance_sheet.html", context)


@login_required
def trial_balance(request):
    """试算平衡表"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.ledger.view', permission_codes):
        messages.error(request, '您没有权限查看试算平衡表')
        return redirect('finance_pages:ledger_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year) if period_year else today.year
        period_month = int(period_month) if period_month else today.month
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # 获取所有会计科目及其余额
    account_subjects = AccountSubject.objects.filter(is_active=True).order_by('code')
    trial_data = []
    
    total_opening_debit = Decimal('0.00')
    total_opening_credit = Decimal('0.00')
    total_period_debit = Decimal('0.00')
    total_period_credit = Decimal('0.00')
    total_closing_debit = Decimal('0.00')
    total_closing_credit = Decimal('0.00')
    
    for subject in account_subjects:
        # 获取该科目的总账记录
        ledger = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        ).order_by('-period_date').first()
        
        if ledger:
            opening_balance = ledger.opening_balance
            period_debit = ledger.period_debit
            period_credit = ledger.period_credit
            closing_balance = ledger.closing_balance
            
            # 根据余额方向调整显示
            if subject.direction == 'debit':
                opening_debit = opening_balance if opening_balance >= 0 else Decimal('0.00')
                opening_credit = -opening_balance if opening_balance < 0 else Decimal('0.00')
                closing_debit = closing_balance if closing_balance >= 0 else Decimal('0.00')
                closing_credit = -closing_balance if closing_balance < 0 else Decimal('0.00')
            else:
                opening_debit = -opening_balance if opening_balance < 0 else Decimal('0.00')
                opening_credit = opening_balance if opening_balance >= 0 else Decimal('0.00')
                closing_debit = -closing_balance if closing_balance < 0 else Decimal('0.00')
                closing_credit = closing_balance if closing_balance >= 0 else Decimal('0.00')
            
            trial_data.append({
                'subject': subject,
                'opening_debit': opening_debit,
                'opening_credit': opening_credit,
                'period_debit': period_debit,
                'period_credit': period_credit,
                'closing_debit': closing_debit,
                'closing_credit': closing_credit,
            })
            
            total_opening_debit += opening_debit
            total_opening_credit += opening_credit
            total_period_debit += period_debit
            total_period_credit += period_credit
            total_closing_debit += closing_debit
            total_closing_credit += closing_credit
    
    # 检查是否平衡
    opening_balanced = abs(total_opening_debit - total_opening_credit) < Decimal('0.01')
    period_balanced = abs(total_period_debit - total_period_credit) < Decimal('0.01')
    closing_balanced = abs(total_closing_debit - total_closing_credit) < Decimal('0.01')
    is_balanced = opening_balanced and period_balanced and closing_balanced
    
    context = _context(
        "试算平衡表",
        "⚖️",
        f"验证 {period_year}年{period_month}月 的借贷是否平衡",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'trial_data': trial_data,
        'period_year': period_year,
        'period_month': period_month,
        'total_opening_debit': total_opening_debit,
        'total_opening_credit': total_opening_credit,
        'total_period_debit': total_period_debit,
        'total_period_credit': total_period_credit,
        'total_closing_debit': total_closing_debit,
        'total_closing_credit': total_closing_credit,
        'opening_balanced': opening_balanced,
        'period_balanced': period_balanced,
        'closing_balanced': closing_balanced,
        'is_balanced': is_balanced,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/trial_balance.html", context)


@login_required
def fund_flow_detail(request, fund_flow_id):
    """资金流水详情"""
    fund_flow = get_object_or_404(
        FundFlow.objects.select_related('project', 'voucher', 'created_by'),
        id=fund_flow_id
    )
    
    context = _context(
        f"资金流水详情 - {fund_flow.flow_number}",
        "💳",
        f"查看资金流水 {fund_flow.flow_number} 的详细信息",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'fund_flow': fund_flow,
    })
    return render(request, "financial_management/fund_flow_detail.html", context)


@login_required
def fund_flow_delete(request, fund_flow_id):
    """删除资金流水"""
    fund_flow = get_object_or_404(FundFlow, id=fund_flow_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.fund_flow.manage', permission_codes):
        messages.error(request, '您没有权限删除资金流水')
        return redirect('finance_pages:fund_flow_detail', fund_flow_id=fund_flow_id)
    
    if request.method == 'POST':
        try:
            flow_number = fund_flow.flow_number
            old_amount = fund_flow.amount
            
            # 删除前先回滚预算使用金额
            if fund_flow.flow_type == 'expense':
                # 创建一个临时对象用于回滚
                temp_flow = FundFlow(
                    flow_type=fund_flow.flow_type,
                    flow_date=fund_flow.flow_date,
                    amount=old_amount,
                    project=fund_flow.project,
                )
                # 回滚：减去旧金额
                _update_budget_from_fund_flow(temp_flow, is_create=False, old_amount=old_amount)
            
            fund_flow.delete()
            messages.success(request, f'资金流水 {flow_number} 已删除')
            return redirect('finance_pages:fund_flow_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除资金流水失败: %s', str(e))
            messages.error(request, f'删除资金流水失败：{str(e)}')
            return redirect('finance_pages:fund_flow_detail', fund_flow_id=fund_flow_id)
    
    context = _context(
        f"删除资金流水 - {fund_flow.flow_number}",
        "🗑️",
        f"确认删除资金流水：{fund_flow.flow_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'fund_flow': fund_flow,
    })
    return render(request, "financial_management/fund_flow_delete.html", context)


# ==================== 财务报表 ====================

@login_required
def report_management(request):
    """财务报表管理"""
    permission_codes = get_user_permission_codes(request.user)
    
    # 获取筛选参数
    report_type = request.GET.get('report_type', '')
    period_year = request.GET.get('period_year', '')
    
    # 获取报表列表
    try:
        reports = FinancialReport.objects.select_related('generated_by').order_by('-report_date', '-generated_time')
        
        if report_type:
            reports = reports.filter(report_type=report_type)
        if period_year:
            reports = reports.filter(period_year=int(period_year))
        
        # 分页
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(reports, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取报表列表失败: %s', str(e))
        page_obj = None
    
    # 统计信息
    try:
        total_reports = FinancialReport.objects.count()
        balance_sheet_count = FinancialReport.objects.filter(report_type='balance_sheet').count()
        income_statement_count = FinancialReport.objects.filter(report_type='income_statement').count()
        cash_flow_count = FinancialReport.objects.filter(report_type='cash_flow').count()
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    context = _context(
        "财务报表管理",
        "📊",
        "管理财务报表生成记录",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'reports': page_obj.object_list if page_obj else [],
        'report_type_choices': FinancialReport.REPORT_TYPE_CHOICES,
        'current_report_type': report_type,
        'current_period_year': period_year,
        'years': range(timezone.now().year - 2, timezone.now().year + 2),
    })
    return render(request, "financial_management/report_list.html", context)


@login_required
def balance_sheet_report(request):
    """资产负债表生成"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.report.view', permission_codes):
        messages.error(request, '您没有权限查看财务报表')
        return redirect('finance_pages:report_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year)
        period_month = int(period_month) if period_month else None
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # 生成资产负债表数据
    report_data = {
        'assets': {},
        'liabilities': {},
        'equity': {},
    }
    
    # 获取资产类科目余额
    asset_subjects = AccountSubject.objects.filter(
        subject_type='asset',
        is_active=True
    ).order_by('code')
    
    total_assets = Decimal('0.00')
    for subject in asset_subjects:
        # 获取该科目的期末余额
        ledger = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        ).order_by('-period_date').first()
        
        if ledger:
            balance = ledger.closing_balance
            if subject.direction == 'credit':
                balance = -balance
            report_data['assets'][subject.code] = {
                'name': subject.name,
                'balance': balance,
            }
            total_assets += balance
    
    # 获取负债类科目余额
    liability_subjects = AccountSubject.objects.filter(
        subject_type='liability',
        is_active=True
    ).order_by('code')
    
    total_liabilities = Decimal('0.00')
    for subject in liability_subjects:
        ledger = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        ).order_by('-period_date').first()
        
        if ledger:
            balance = ledger.closing_balance
            if subject.direction == 'debit':
                balance = -balance
            report_data['liabilities'][subject.code] = {
                'name': subject.name,
                'balance': balance,
            }
            total_liabilities += balance
    
    # 获取所有者权益类科目余额
    equity_subjects = AccountSubject.objects.filter(
        subject_type='equity',
        is_active=True
    ).order_by('code')
    
    total_equity = Decimal('0.00')
    for subject in equity_subjects:
        ledger = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        ).order_by('-period_date').first()
        
        if ledger:
            balance = ledger.closing_balance
            if subject.direction == 'debit':
                balance = -balance
            report_data['equity'][subject.code] = {
                'name': subject.name,
                'balance': balance,
            }
            total_equity += balance
    
    report_data['total_assets'] = total_assets
    report_data['total_liabilities'] = total_liabilities
    report_data['total_equity'] = total_equity
    report_data['total_liabilities_equity'] = total_liabilities + total_equity
    
    # 如果请求生成报表，保存报表记录
    if request.method == 'POST':
        try:
            # 生成唯一的报表编号
            base_number = f'BS-{period_year}-{period_month:02d}' if period_month else f'BS-{period_year}'
            report_number = base_number
            counter = 1
            while FinancialReport.objects.filter(report_number=report_number).exists():
                report_number = f'{base_number}-{counter:02d}'
                counter += 1
            
            report = FinancialReport.objects.create(
                report_number=report_number,
                report_type='balance_sheet',
                period_year=period_year,
                period_month=period_month,
                report_date=today,
                report_data=report_data,
                generated_by=request.user,
            )
            messages.success(request, f'资产负债表生成成功！报表编号：{report_number}')
            return redirect('finance_pages:report_detail', report_id=report.id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('生成资产负债表失败: %s', str(e))
            messages.error(request, f'生成报表失败：{str(e)}')
    
    context = _context(
        "资产负债表",
        "📊",
        f"{period_year}年{period_month}月资产负债表",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'report_data': report_data,
        'period_year': period_year,
        'period_month': period_month,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/balance_sheet_report.html", context)


@login_required
def income_statement_report(request):
    """利润表生成"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.report.view', permission_codes):
        messages.error(request, '您没有权限查看财务报表')
        return redirect('finance_pages:report_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year)
        period_month = int(period_month) if period_month else None
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # 生成利润表数据
    report_data = {
        'revenue': {},
        'expenses': {},
        'costs': {},
    }
    
    # 获取收入类科目发生额
    revenue_subjects = AccountSubject.objects.filter(
        subject_type='revenue',
        is_active=True
    ).order_by('code')
    
    total_revenue = Decimal('0.00')
    for subject in revenue_subjects:
        ledgers = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        )
        period_credit = sum(l.period_credit for l in ledgers)
        report_data['revenue'][subject.code] = {
            'name': subject.name,
            'amount': period_credit,
        }
        total_revenue += period_credit
    
    # 获取费用类科目发生额
    expense_subjects = AccountSubject.objects.filter(
        subject_type='expense',
        is_active=True
    ).order_by('code')
    
    total_expenses = Decimal('0.00')
    for subject in expense_subjects:
        ledgers = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        )
        period_debit = sum(l.period_debit for l in ledgers)
        report_data['expenses'][subject.code] = {
            'name': subject.name,
            'amount': period_debit,
        }
        total_expenses += period_debit
    
    # 获取成本类科目发生额
    cost_subjects = AccountSubject.objects.filter(
        subject_type='cost',
        is_active=True
    ).order_by('code')
    
    total_costs = Decimal('0.00')
    for subject in cost_subjects:
        ledgers = Ledger.objects.filter(
            account_subject=subject,
            period_year=period_year,
            period_month=period_month
        )
        period_debit = sum(l.period_debit for l in ledgers)
        report_data['costs'][subject.code] = {
            'name': subject.name,
            'amount': period_debit,
        }
        total_costs += period_debit
    
    report_data['total_revenue'] = total_revenue
    report_data['total_costs'] = total_costs
    report_data['total_expenses'] = total_expenses
    report_data['gross_profit'] = total_revenue - total_costs
    report_data['net_profit'] = total_revenue - total_costs - total_expenses
    
    # 如果请求生成报表，保存报表记录
    if request.method == 'POST':
        try:
            # 生成唯一的报表编号
            base_number = f'IS-{period_year}-{period_month:02d}' if period_month else f'IS-{period_year}'
            report_number = base_number
            counter = 1
            while FinancialReport.objects.filter(report_number=report_number).exists():
                report_number = f'{base_number}-{counter:02d}'
                counter += 1
            
            report = FinancialReport.objects.create(
                report_number=report_number,
                report_type='income_statement',
                period_year=period_year,
                period_month=period_month,
                report_date=today,
                report_data=report_data,
                generated_by=request.user,
            )
            messages.success(request, f'利润表生成成功！报表编号：{report_number}')
            return redirect('finance_pages:report_detail', report_id=report.id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('生成利润表失败: %s', str(e))
            messages.error(request, f'生成报表失败：{str(e)}')
    
    context = _context(
        "利润表",
        "📈",
        f"{period_year}年{period_month}月利润表",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'report_data': report_data,
        'period_year': period_year,
        'period_month': period_month,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/income_statement_report.html", context)


@login_required
def cash_flow_report(request):
    """现金流量表生成"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.report.view', permission_codes):
        messages.error(request, '您没有权限查看财务报表')
        return redirect('finance_pages:report_management')
    
    today = timezone.now().date()
    period_year = request.GET.get('period_year', str(today.year))
    period_month = request.GET.get('period_month', str(today.month))
    
    try:
        period_year = int(period_year)
        period_month = int(period_month) if period_month else None
    except (ValueError, TypeError):
        period_year = today.year
        period_month = today.month
    
    # 生成现金流量表数据（基于资金流水）
    if period_month:
        start_date = today.replace(year=period_year, month=period_month, day=1)
        if period_month == 12:
            end_date = today.replace(year=period_year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_date = today.replace(year=period_year, month=period_month + 1, day=1) - timedelta(days=1)
    else:
        start_date = today.replace(year=period_year, month=1, day=1)
        end_date = today.replace(year=period_year, month=12, day=31)
    
    # 经营活动现金流量
    operating_income = FundFlow.objects.filter(
        flow_date__gte=start_date,
        flow_date__lte=end_date,
        flow_type='income'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    operating_expense = FundFlow.objects.filter(
        flow_date__gte=start_date,
        flow_date__lte=end_date,
        flow_type='expense'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    operating_cash_flow = operating_income - operating_expense
    
    # 投资活动现金流量（简化处理）
    investing_cash_flow = Decimal('0.00')
    
    # 筹资活动现金流量（简化处理）
    financing_cash_flow = Decimal('0.00')
    
    report_data = {
        'operating': {
            'income': operating_income,
            'expense': operating_expense,
            'net': operating_cash_flow,
        },
        'investing': {
            'net': investing_cash_flow,
        },
        'financing': {
            'net': financing_cash_flow,
        },
        'net_cash_flow': operating_cash_flow + investing_cash_flow + financing_cash_flow,
    }
    
    # 如果请求生成报表，保存报表记录
    if request.method == 'POST':
        try:
            # 生成唯一的报表编号
            base_number = f'CF-{period_year}-{period_month:02d}' if period_month else f'CF-{period_year}'
            report_number = base_number
            counter = 1
            while FinancialReport.objects.filter(report_number=report_number).exists():
                report_number = f'{base_number}-{counter:02d}'
                counter += 1
            
            report = FinancialReport.objects.create(
                report_number=report_number,
                report_type='cash_flow',
                period_year=period_year,
                period_month=period_month,
                report_date=today,
                report_data=report_data,
                generated_by=request.user,
            )
            messages.success(request, f'现金流量表生成成功！报表编号：{report_number}')
            return redirect('finance_pages:report_detail', report_id=report.id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('生成现金流量表失败: %s', str(e))
            messages.error(request, f'生成报表失败：{str(e)}')
    
    context = _context(
        "现金流量表",
        "💳",
        f"{period_year}年{period_month}月现金流量表",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'report_data': report_data,
        'period_year': period_year,
        'period_month': period_month,
        'years': range(today.year - 2, today.year + 2),
        'months': range(1, 13),
    })
    return render(request, "financial_management/cash_flow_report.html", context)


@login_required
def report_export(request, report_id):
    """导出财务报表为Excel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.report.view', permission_codes):
        messages.error(request, '您没有权限导出财务报表')
        return redirect('finance_pages:report_management')
    
    report = get_object_or_404(FinancialReport, id=report_id)
    report_data = report.report_data or {}
    
    # 创建Excel工作簿
    workbook = Workbook()
    worksheet = workbook.active
    
    # 根据报表类型设置标题和内容
    report_type_dict = dict(FinancialReport.TYPE_CHOICES)
    worksheet.title = report_type_dict.get(report.report_type, '财务报表')
    
    # 设置报表标题
    title = f"{report_type_dict.get(report.report_type, '财务报表')} - {report.period_year}年"
    if report.period_month:
        title += f"{report.period_month}月"
    worksheet.append([title])
    
    # 合并标题单元格
    worksheet.merge_cells(f'A1:{get_column_letter(10)}1')
    title_cell = worksheet['A1']
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    worksheet.append([])  # 空行
    
    # 根据报表类型导出数据
    if report.report_type == 'balance_sheet':
        # 资产负债表
        headers = ['项目', '期初余额', '本期借方', '本期贷方', '期末余额']
        worksheet.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加数据
        for item in report_data.get('items', []):
            row = [
                item.get('name', ''),
                float(item.get('opening_balance', 0)),
                float(item.get('period_debit', 0)),
                float(item.get('period_credit', 0)),
                float(item.get('closing_balance', 0)),
            ]
            worksheet.append(row)
    
    elif report.report_type == 'income_statement':
        # 利润表
        headers = ['项目', '金额']
        worksheet.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加收入数据
        worksheet.append(['收入'])
        for code, data in report_data.get('revenue', {}).items():
            worksheet.append(['', data.get('name', ''), float(data.get('amount', 0))])
        worksheet.append(['收入合计', '', float(report_data.get('total_revenue', 0))])
        worksheet.append([])
        
        # 添加成本数据
        worksheet.append(['成本'])
        for code, data in report_data.get('costs', {}).items():
            worksheet.append(['', data.get('name', ''), float(data.get('amount', 0))])
        worksheet.append(['成本合计', '', float(report_data.get('total_costs', 0))])
        worksheet.append([])
        
        # 添加费用数据
        worksheet.append(['费用'])
        for code, data in report_data.get('expenses', {}).items():
            worksheet.append(['', data.get('name', ''), float(data.get('amount', 0))])
        worksheet.append(['费用合计', '', float(report_data.get('total_expenses', 0))])
        worksheet.append([])
        
        # 添加利润数据
        worksheet.append(['毛利润', '', float(report_data.get('gross_profit', 0))])
        worksheet.append(['净利润', '', float(report_data.get('net_profit', 0))])
    
    elif report.report_type == 'cash_flow':
        # 现金流量表
        headers = ['项目', '金额']
        worksheet.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 经营活动现金流量
        operating = report_data.get('operating', {})
        worksheet.append(['经营活动现金流量'])
        worksheet.append(['', '现金流入', float(operating.get('income', 0))])
        worksheet.append(['', '现金流出', float(operating.get('expense', 0))])
        worksheet.append(['', '净流量', float(operating.get('net', 0))])
        worksheet.append([])
        
        # 投资活动现金流量
        investing = report_data.get('investing', {})
        worksheet.append(['投资活动现金流量'])
        worksheet.append(['', '净流量', float(investing.get('net', 0))])
        worksheet.append([])
        
        # 筹资活动现金流量
        financing = report_data.get('financing', {})
        worksheet.append(['筹资活动现金流量'])
        worksheet.append(['', '净流量', float(financing.get('net', 0))])
        worksheet.append([])
        
        # 现金净流量
        worksheet.append(['现金净流量', '', float(report_data.get('net_cash_flow', 0))])
    
    # 调整列宽
    column_widths = [30, 15, 15, 15, 15]
    for i, width in enumerate(column_widths[:len(headers)], 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width
    
    # 添加报表信息
    worksheet.append([])
    worksheet.append(['报表编号', report.report_number])
    worksheet.append(['生成时间', report.generated_time.strftime('%Y-%m-%d %H:%M:%S') if report.generated_time else ''])
    worksheet.append(['生成人', report.generated_by.get_full_name() if report.generated_by else ''])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'{report.report_number}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def report_detail(request, report_id):
    """报表详情"""
    report = get_object_or_404(FinancialReport.objects.select_related('generated_by'), id=report_id)
    
    context = _context(
        f"报表详情 - {report.report_number}",
        "📊",
        f"查看{report.get_report_type_display()}详情",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'report': report,
    })
    return render(request, "financial_management/report_detail.html", context)


# ==================== 往来账款 ====================

@login_required
def receivable_management(request):
    """应收账款管理"""
    permission_codes = get_user_permission_codes(request.user)
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # 获取应收账款列表
    try:
        receivables = ReceivableAccount.objects.select_related('customer', 'project', 'created_by').order_by('-receivable_date', '-account_number')
        
        if search:
            receivables = receivables.filter(
                Q(account_number__icontains=search) |
                Q(description__icontains=search)
            )
        if status:
            receivables = receivables.filter(status=status)
        if date_from:
            receivables = receivables.filter(receivable_date__gte=date_from)
        if date_to:
            receivables = receivables.filter(receivable_date__lte=date_to)
        
        # 分页
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(receivables, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取应收账款列表失败: %s', str(e))
        page_obj = None
    
    # 统计信息
    try:
        total_receivables = ReceivableAccount.objects.count()
        total_amount = ReceivableAccount.objects.aggregate(total=Sum('receivable_amount'))['total'] or Decimal('0')
        received_amount = ReceivableAccount.objects.aggregate(total=Sum('received_amount'))['total'] or Decimal('0')
        remaining_amount = ReceivableAccount.objects.aggregate(total=Sum('remaining_amount'))['total'] or Decimal('0')
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    context = _context(
        "应收账款管理",
        "💰",
        "管理应收账款记录",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'receivables': page_obj.object_list if page_obj else [],
        'status_choices': ReceivableAccount.STATUS_CHOICES,
        'current_search': search,
        'current_status': status,
        'current_date_from': date_from,
        'current_date_to': date_to,
    })
    return render(request, "financial_management/receivable_list.html", context)


@login_required
def receivable_export(request):
    """导出应收账款列表为Excel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.receivable.view', permission_codes):
        messages.error(request, '您没有权限导出应收账款')
        return redirect('finance_pages:receivable_management')
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    receivables = ReceivableAccount.objects.select_related('project', 'created_by').order_by('-receivable_date', '-account_number')
    
    if search:
        receivables = receivables.filter(
            Q(account_number__icontains=search) |
            Q(customer__icontains=search) |
            Q(description__icontains=search)
        )
    if status:
        receivables = receivables.filter(status=status)
    if date_from:
        receivables = receivables.filter(receivable_date__gte=date_from)
    if date_to:
        receivables = receivables.filter(receivable_date__lte=date_to)
    
    # 创建Excel工作簿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '应收账款'
    
    headers = ['应收单号', '应收日期', '客户名称', '应收金额', '已收金额', '未收金额', '到期日期', '账期(天)', '状态', '关联项目', '备注', '创建人', '创建时间']
    worksheet.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加数据
    status_dict = dict(ReceivableAccount.STATUS_CHOICES)
    for receivable in receivables:
        row = [
            receivable.account_number,
            receivable.receivable_date.strftime('%Y-%m-%d') if receivable.receivable_date else '',
            receivable.customer,
            float(receivable.receivable_amount),
            float(receivable.paid_amount),
            float(receivable.remaining_amount),
            receivable.due_date.strftime('%Y-%m-%d') if receivable.due_date else '',
            receivable.payment_terms or '',
            status_dict.get(receivable.status, receivable.status),
            receivable.project.project_number if receivable.project else '',
            receivable.description or '',
            receivable.created_by.get_full_name() if receivable.created_by else '',
            receivable.created_time.strftime('%Y-%m-%d %H:%M') if receivable.created_time else '',
        ]
        worksheet.append(row)
    
    # 调整列宽
    column_widths = [18, 12, 20, 12, 12, 12, 12, 10, 10, 15, 30, 12, 18]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('应收账款_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def receivable_create(request):
    """新增应收账款"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.receivable.create', permission_codes):
        messages.error(request, '您没有权限创建应收账款')
        return redirect('finance_pages:receivable_management')
    
    if request.method == 'POST':
        from .forms import ReceivableAccountForm
        form = ReceivableAccountForm(request.POST)
        if form.is_valid():
            receivable = form.save(commit=False)
            # 自动生成应收单号
            if not receivable.account_number:
                current_year = timezone.now().year
                # 查找当前年度最大的序号
                max_receivable = ReceivableAccount.objects.filter(
                    account_number__startswith=f'AR-{current_year}-'
                ).order_by('-account_number').first()
                
                if max_receivable:
                    try:
                        # 提取序号部分
                        parts = max_receivable.account_number.split('-')
                        if len(parts) >= 3:
                            seq = int(parts[-1]) + 1
                        else:
                            seq = 1
                    except (ValueError, IndexError):
                        seq = 1
                else:
                    seq = 1
                receivable.account_number = f'AR-{current_year}-{seq:04d}'
            
            # 如果设置了应收日期和账期，自动计算到期日期
            if receivable.receivable_date and receivable.payment_terms and not receivable.due_date:
                receivable.due_date = receivable.receivable_date + timedelta(days=receivable.payment_terms)
            
            receivable.created_by = request.user
            receivable.save()
            messages.success(request, f'应收账款 {receivable.account_number} 创建成功！')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable.id)
    else:
        from .forms import ReceivableAccountForm
        form = ReceivableAccountForm()
        form.fields['receivable_date'].initial = timezone.now().date()
    
    context = _context(
        "新增应收账款",
        "➕",
        "创建新的应收账款记录",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/receivable_form.html", context)


@login_required
def receivable_update(request, receivable_id):
    """编辑应收账款"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.receivable.manage', permission_codes):
        messages.error(request, '您没有权限编辑应收账款')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    receivable = get_object_or_404(ReceivableAccount, id=receivable_id)
    
    if request.method == 'POST':
        from .forms import ReceivableAccountForm
        form = ReceivableAccountForm(request.POST, instance=receivable)
        if form.is_valid():
            receivable = form.save(commit=False)
            # 如果设置了应收日期和账期，自动计算到期日期
            if receivable.receivable_date and receivable.payment_terms and not receivable.due_date:
                receivable.due_date = receivable.receivable_date + timedelta(days=receivable.payment_terms)
            receivable.save()
            messages.success(request, f'应收账款 {receivable.account_number} 更新成功！')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable.id)
    else:
        from .forms import ReceivableAccountForm
        form = ReceivableAccountForm(instance=receivable)
    
    context = _context(
        f"编辑应收账款 - {receivable.account_number}",
        "✏️",
        f"编辑应收账款 {receivable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'receivable': receivable,
        'is_create': False,
    })
    return render(request, "financial_management/receivable_form.html", context)


@login_required
def receivable_detail(request, receivable_id):
    """应收账款详情"""
    receivable = get_object_or_404(
        ReceivableAccount.objects.select_related('customer', 'project', 'created_by'),
        id=receivable_id
    )
    
    # 获取收款历史记录（通过资金流水）
    payment_history = FundFlow.objects.filter(
        flow_type='income',
        summary__icontains=receivable.account_number
    ).select_related('created_by', 'project').order_by('-flow_date', '-created_time')
    
    # 也可以根据对方单位匹配
    if receivable.customer:
        payment_history = payment_history.filter(
            Q(summary__icontains=receivable.account_number) |
            Q(counterparty__icontains=receivable.customer.name)
        )
    
    context = _context(
        f"应收账款详情 - {receivable.account_number}",
        "💰",
        f"查看应收账款 {receivable.account_number} 的详细信息",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'receivable': receivable,
        'payment_history': payment_history[:20],  # 最近20条记录
        'payment_history_count': payment_history.count(),
    })
    return render(request, "financial_management/receivable_detail.html", context)


@login_required
def receivable_payment(request, receivable_id):
    """记录应收账款收款"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.receivable.manage', permission_codes):
        messages.error(request, '您没有权限记录收款')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    receivable = get_object_or_404(ReceivableAccount, id=receivable_id)
    
    if receivable.status == 'completed':
        messages.error(request, '该应收账款已完成收款，不能继续收款')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    if request.method == 'POST':
        payment_amount_str = request.POST.get('payment_amount', '0')
        payment_date = request.POST.get('payment_date', '')
        payment_method = request.POST.get('payment_method', '')
        payment_notes = request.POST.get('payment_notes', '')
        
        try:
            payment_amount = Decimal(payment_amount_str)
            if payment_amount <= 0:
                messages.error(request, '收款金额必须大于0')
                return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
            
            if payment_amount > receivable.remaining_amount:
                messages.error(request, f'收款金额不能超过未收金额 {receivable.remaining_amount:,.2f}')
                return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
            
            # 更新应收账款
            receivable.received_amount += payment_amount
            receivable.remaining_amount = receivable.receivable_amount - receivable.received_amount
            
            # 自动更新状态
            if receivable.remaining_amount <= 0:
                receivable.status = 'completed'
            elif receivable.received_amount > 0:
                receivable.status = 'partial'
            
            receivable.save()
            
            # 可选：创建资金流水记录
            try:
                from django.db import transaction
                with transaction.atomic():
                    current_year = timezone.now().year
                    max_flow = FundFlow.objects.filter(
                        flow_number__startswith=f'FLOW-{current_year}-'
                    ).order_by('-flow_number').first()
                    
                    if max_flow:
                        try:
                            seq = int(max_flow.flow_number.split('-')[-1]) + 1
                        except (ValueError, IndexError):
                            seq = 1
                    else:
                        seq = 1
                    
                    flow_number = f'FLOW-{current_year}-{seq:04d}'
                    
                    FundFlow.objects.create(
                        flow_number=flow_number,
                        flow_date=payment_date or timezone.now().date(),
                        flow_type='income',
                        amount=payment_amount,
                        account_name=payment_method or '银行账户',
                        counterparty=receivable.customer.name if receivable.customer else '',
                        summary=f'应收账款收款：{receivable.account_number}' + (f' - {payment_notes}' if payment_notes else ''),
                        project=receivable.project,
                        created_by=request.user,
                    )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.exception('创建资金流水失败: %s', str(e))
                # 不影响收款记录，只记录日志
            
            messages.success(request, f'成功记录收款 {payment_amount:,.2f} 元')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable.id)
            
        except (ValueError, InvalidOperation):
            messages.error(request, '收款金额格式错误')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('记录收款失败: %s', str(e))
            messages.error(request, f'记录收款失败：{str(e)}')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    # GET请求，显示收款表单
    context = _context(
        f"记录收款 - {receivable.account_number}",
        "💰",
        f"记录应收账款 {receivable.account_number} 的收款",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'receivable': receivable,
    })
    return render(request, "financial_management/receivable_payment.html", context)


@login_required
def receivable_cancel(request, receivable_id):
    """取消应收账款"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.receivable.manage', permission_codes):
        messages.error(request, '您没有权限取消应收账款')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    receivable = get_object_or_404(ReceivableAccount, id=receivable_id)
    
    # 检查状态：已完成的应收账款不能取消
    if receivable.status == 'completed':
        messages.error(request, '已完成的应收账款不能取消')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable.id)
    
    if request.method == 'POST':
        receivable.status = 'cancelled'
        receivable.save()
        messages.success(request, f'应收账款 {receivable.account_number} 已取消')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable.id)
    
    # GET请求，显示确认页面
    context = _context(
        f"取消应收账款 - {receivable.account_number}",
        "❌",
        f"取消应收账款 {receivable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'receivable': receivable,
    })
    return render(request, "financial_management/receivable_cancel.html", context)


@login_required
def receivable_delete(request, receivable_id):
    """删除应收账款"""
    receivable = get_object_or_404(ReceivableAccount, id=receivable_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.receivable.manage', permission_codes):
        messages.error(request, '您没有权限删除应收账款')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    # 如果已收款，不允许删除
    if receivable.received_amount > 0:
        messages.error(request, '该应收账款已有收款记录，无法删除')
        return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    if request.method == 'POST':
        try:
            account_number = receivable.account_number
            receivable.delete()
            messages.success(request, f'应收账款 {account_number} 已删除')
            return redirect('finance_pages:receivable_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除应收账款失败: %s', str(e))
            messages.error(request, f'删除应收账款失败：{str(e)}')
            return redirect('finance_pages:receivable_detail', receivable_id=receivable_id)
    
    context = _context(
        f"删除应收账款 - {receivable.account_number}",
        "🗑️",
        f"确认删除应收账款：{receivable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'receivable': receivable,
    })
    return render(request, "financial_management/receivable_delete.html", context)


@login_required
def payable_management(request):
    """应付账款管理"""
    permission_codes = get_user_permission_codes(request.user)
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # 获取应付账款列表
    try:
        payables = PayableAccount.objects.select_related('project', 'created_by').order_by('-payable_date', '-account_number')
        
        if search:
            payables = payables.filter(
                Q(account_number__icontains=search) |
                Q(supplier__icontains=search) |
                Q(description__icontains=search)
            )
        if status:
            payables = payables.filter(status=status)
        if date_from:
            payables = payables.filter(payable_date__gte=date_from)
        if date_to:
            payables = payables.filter(payable_date__lte=date_to)
        
        # 分页
        page_size = request.GET.get('page_size', '10')
        try:
            per_page = int(page_size)
            if per_page not in [10, 20, 50]:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        paginator = Paginator(payables, per_page)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取应付账款列表失败: %s', str(e))
        page_obj = None
    
    # 统计信息
    try:
        total_payables = PayableAccount.objects.count()
        total_amount = PayableAccount.objects.aggregate(total=Sum('payable_amount'))['total'] or Decimal('0')
        paid_amount = PayableAccount.objects.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
        remaining_amount = PayableAccount.objects.aggregate(total=Sum('remaining_amount'))['total'] or Decimal('0')
        
        summary_cards = []
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception('获取统计信息失败: %s', str(e))
        summary_cards = []
    
    context = _context(
        "应付账款管理",
        "💸",
        "管理应付账款记录",
        summary_cards=summary_cards,
        request=request,
        use_financial_nav=True
    )
    context.update({
        'page_obj': page_obj,
        'payables': page_obj.object_list if page_obj else [],
        'status_choices': PayableAccount.STATUS_CHOICES,
        'current_search': search,
        'current_status': status,
        'current_date_from': date_from,
        'current_date_to': date_to,
    })
    return render(request, "financial_management/payable_list.html", context)


@login_required
def payable_export(request):
    """导出应付账款列表为Excel"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.payable.view', permission_codes):
        messages.error(request, '您没有权限导出应付账款')
        return redirect('finance_pages:payable_management')
    
    # 获取筛选参数
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    payables = PayableAccount.objects.select_related('project', 'created_by').order_by('-payable_date', '-account_number')
    
    if search:
        payables = payables.filter(
            Q(account_number__icontains=search) |
            Q(supplier__icontains=search) |
            Q(description__icontains=search)
        )
    if status:
        payables = payables.filter(status=status)
    if date_from:
        payables = payables.filter(payable_date__gte=date_from)
    if date_to:
        payables = payables.filter(payable_date__lte=date_to)
    
    # 创建Excel工作簿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '应付账款'
    
    headers = ['应付单号', '应付日期', '供应商', '应付金额', '已付金额', '未付金额', '到期日期', '账期(天)', '状态', '关联项目', '备注', '创建人', '创建时间']
    worksheet.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 添加数据
    status_dict = dict(PayableAccount.STATUS_CHOICES)
    for payable in payables:
        row = [
            payable.account_number,
            payable.payable_date.strftime('%Y-%m-%d') if payable.payable_date else '',
            payable.supplier,
            float(payable.payable_amount),
            float(payable.paid_amount),
            float(payable.remaining_amount),
            payable.due_date.strftime('%Y-%m-%d') if payable.due_date else '',
            payable.payment_terms or '',
            status_dict.get(payable.status, payable.status),
            payable.project.project_number if payable.project else '',
            payable.description or '',
            payable.created_by.get_full_name() if payable.created_by else '',
            payable.created_time.strftime('%Y-%m-%d %H:%M') if payable.created_time else '',
        ]
        worksheet.append(row)
    
    # 调整列宽
    column_widths = [18, 12, 20, 12, 12, 12, 12, 10, 10, 15, 30, 12, 18]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = timezone.now().strftime('应付账款_%Y%m%d_%H%M%S.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
def payable_create(request):
    """新增应付账款"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.payable.create', permission_codes):
        messages.error(request, '您没有权限创建应付账款')
        return redirect('finance_pages:payable_management')
    
    if request.method == 'POST':
        from .forms import PayableAccountForm
        form = PayableAccountForm(request.POST)
        if form.is_valid():
            payable = form.save(commit=False)
            # 自动生成应付单号
            if not payable.account_number:
                current_year = timezone.now().year
                # 查找当前年度最大的序号
                max_payable = PayableAccount.objects.filter(
                    account_number__startswith=f'AP-{current_year}-'
                ).order_by('-account_number').first()
                
                if max_payable:
                    try:
                        # 提取序号部分
                        parts = max_payable.account_number.split('-')
                        if len(parts) >= 3:
                            seq = int(parts[-1]) + 1
                        else:
                            seq = 1
                    except (ValueError, IndexError):
                        seq = 1
                else:
                    seq = 1
                payable.account_number = f'AP-{current_year}-{seq:04d}'
            
            # 如果设置了应付日期和账期，自动计算到期日期
            if payable.payable_date and payable.payment_terms and not payable.due_date:
                payable.due_date = payable.payable_date + timedelta(days=payable.payment_terms)
            
            payable.created_by = request.user
            payable.save()
            messages.success(request, f'应付账款 {payable.account_number} 创建成功！')
            return redirect('finance_pages:payable_detail', payable_id=payable.id)
    else:
        from .forms import PayableAccountForm
        form = PayableAccountForm()
        form.fields['payable_date'].initial = timezone.now().date()
    
    context = _context(
        "新增应付账款",
        "➕",
        "创建新的应付账款记录",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'is_create': True,
    })
    return render(request, "financial_management/payable_form.html", context)


@login_required
def payable_update(request, payable_id):
    """编辑应付账款"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.payable.manage', permission_codes):
        messages.error(request, '您没有权限编辑应付账款')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    payable = get_object_or_404(PayableAccount, id=payable_id)
    
    if request.method == 'POST':
        from .forms import PayableAccountForm
        form = PayableAccountForm(request.POST, instance=payable)
        if form.is_valid():
            payable = form.save(commit=False)
            # 如果设置了应付日期和账期，自动计算到期日期
            if payable.payable_date and payable.payment_terms and not payable.due_date:
                payable.due_date = payable.payable_date + timedelta(days=payable.payment_terms)
            payable.save()
            messages.success(request, f'应付账款 {payable.account_number} 更新成功！')
            return redirect('finance_pages:payable_detail', payable_id=payable.id)
    else:
        from .forms import PayableAccountForm
        form = PayableAccountForm(instance=payable)
    
    context = _context(
        f"编辑应付账款 - {payable.account_number}",
        "✏️",
        f"编辑应付账款 {payable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'form': form,
        'payable': payable,
        'is_create': False,
    })
    return render(request, "financial_management/payable_form.html", context)


@login_required
def payable_detail(request, payable_id):
    """应付账款详情"""
    payable = get_object_or_404(
        PayableAccount.objects.select_related('project', 'created_by'),
        id=payable_id
    )
    
    # 获取付款历史记录（通过资金流水）
    payment_history = FundFlow.objects.filter(
        flow_type='expense',
        summary__icontains=payable.account_number
    ).select_related('created_by', 'project').order_by('-flow_date', '-created_time')
    
    # 也可以根据供应商匹配
    payment_history = payment_history.filter(
        Q(summary__icontains=payable.account_number) |
        Q(counterparty__icontains=payable.supplier)
    )
    
    context = _context(
        f"应付账款详情 - {payable.account_number}",
        "💸",
        f"查看应付账款 {payable.account_number} 的详细信息",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'payable': payable,
        'payment_history': payment_history[:20],  # 最近20条记录
        'payment_history_count': payment_history.count(),
    })
    return render(request, "financial_management/payable_detail.html", context)


@login_required
def payable_payment(request, payable_id):
    """记录应付账款付款"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.payable.manage', permission_codes):
        messages.error(request, '您没有权限记录付款')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    payable = get_object_or_404(PayableAccount, id=payable_id)
    
    if payable.status == 'completed':
        messages.error(request, '该应付账款已完成付款，不能继续付款')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    if request.method == 'POST':
        payment_amount_str = request.POST.get('payment_amount', '0')
        payment_date = request.POST.get('payment_date', '')
        payment_method = request.POST.get('payment_method', '')
        payment_notes = request.POST.get('payment_notes', '')
        
        try:
            payment_amount = Decimal(payment_amount_str)
            if payment_amount <= 0:
                messages.error(request, '付款金额必须大于0')
                return redirect('finance_pages:payable_detail', payable_id=payable_id)
            
            if payment_amount > payable.remaining_amount:
                messages.error(request, f'付款金额不能超过未付金额 {payable.remaining_amount:,.2f}')
                return redirect('finance_pages:payable_detail', payable_id=payable_id)
            
            # 更新应付账款
            payable.paid_amount += payment_amount
            payable.remaining_amount = payable.payable_amount - payable.paid_amount
            
            # 自动更新状态
            if payable.remaining_amount <= 0:
                payable.status = 'completed'
            elif payable.paid_amount > 0:
                payable.status = 'partial'
            
            payable.save()
            
            # 可选：创建资金流水记录
            try:
                from django.db import transaction
                with transaction.atomic():
                    current_year = timezone.now().year
                    max_flow = FundFlow.objects.filter(
                        flow_number__startswith=f'FLOW-{current_year}-'
                    ).order_by('-flow_number').first()
                    
                    if max_flow:
                        try:
                            seq = int(max_flow.flow_number.split('-')[-1]) + 1
                        except (ValueError, IndexError):
                            seq = 1
                    else:
                        seq = 1
                    
                    flow_number = f'FLOW-{current_year}-{seq:04d}'
                    
                    FundFlow.objects.create(
                        flow_number=flow_number,
                        flow_date=payment_date or timezone.now().date(),
                        flow_type='expense',
                        amount=payment_amount,
                        account_name=payment_method or '银行账户',
                        counterparty=payable.supplier,
                        summary=f'应付账款付款：{payable.account_number}' + (f' - {payment_notes}' if payment_notes else ''),
                        project=payable.project,
                        created_by=request.user,
                    )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.exception('创建资金流水失败: %s', str(e))
                # 不影响付款记录，只记录日志
            
            messages.success(request, f'成功记录付款 {payment_amount:,.2f} 元')
            return redirect('finance_pages:payable_detail', payable_id=payable.id)
            
        except (ValueError, InvalidOperation):
            messages.error(request, '付款金额格式错误')
            return redirect('finance_pages:payable_detail', payable_id=payable_id)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('记录付款失败: %s', str(e))
            messages.error(request, f'记录付款失败：{str(e)}')
            return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    # GET请求，显示付款表单
    context = _context(
        f"记录付款 - {payable.account_number}",
        "💸",
        f"记录应付账款 {payable.account_number} 的付款",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'payable': payable,
    })
    return render(request, "financial_management/payable_payment.html", context)


@login_required
def payable_cancel(request, payable_id):
    """取消应付账款"""
    permission_codes = get_user_permission_codes(request.user)
    if not _permission_granted('financial_management.payable.manage', permission_codes):
        messages.error(request, '您没有权限取消应付账款')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    payable = get_object_or_404(PayableAccount, id=payable_id)
    
    # 检查状态：已完成的应付账款不能取消
    if payable.status == 'completed':
        messages.error(request, '已完成的应付账款不能取消')
        return redirect('finance_pages:payable_detail', payable_id=payable.id)
    
    if request.method == 'POST':
        payable.status = 'cancelled'
        payable.save()
        messages.success(request, f'应付账款 {payable.account_number} 已取消')
        return redirect('finance_pages:payable_detail', payable_id=payable.id)
    
    # GET请求，显示确认页面
    context = _context(
        f"取消应付账款 - {payable.account_number}",
        "❌",
        f"取消应付账款 {payable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'payable': payable,
    })
    return render(request, "financial_management/payable_cancel.html", context)


@login_required
def payable_delete(request, payable_id):
    """删除应付账款"""
    payable = get_object_or_404(PayableAccount, id=payable_id)
    permission_codes = get_user_permission_codes(request.user)
    
    if not _permission_granted('financial_management.payable.manage', permission_codes):
        messages.error(request, '您没有权限删除应付账款')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    # 如果已付款，不允许删除
    if payable.paid_amount > 0:
        messages.error(request, '该应付账款已有付款记录，无法删除')
        return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    if request.method == 'POST':
        try:
            account_number = payable.account_number
            payable.delete()
            messages.success(request, f'应付账款 {account_number} 已删除')
            return redirect('finance_pages:payable_management')
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception('删除应付账款失败: %s', str(e))
            messages.error(request, f'删除应付账款失败：{str(e)}')
            return redirect('finance_pages:payable_detail', payable_id=payable_id)
    
    context = _context(
        f"删除应付账款 - {payable.account_number}",
        "🗑️",
        f"确认删除应付账款：{payable.account_number}",
        request=request,
        use_financial_nav=True
    )
    context.update({
        'payable': payable,
    })
    return render(request, "financial_management/payable_delete.html", context)

